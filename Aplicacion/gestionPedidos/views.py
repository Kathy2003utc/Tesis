import json
import re
from .utils.push import enviar_push
from django.db import transaction, IntegrityError
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from Restaurante.settings import BASE_DIR
from .decorators import rol_requerido
from django.db.models import Prefetch
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import check_password
from .models import Usuario, Mesa, Pedido, DetallePedido, Producto, Notificacion, Mensaje, Pago, Comprobante, Horario, ComprobanteCliente, PushSubscription
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum, F
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from django.views.decorators.csrf import csrf_protect
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import localtime
from django.db.models import Exists, OuterRef
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP
from django.db import transaction
from django.http import HttpResponse
from django.template.loader import get_template
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Image
import os
from django.utils.dateparse import parse_date
from django.db.models import Q
from django.db import IntegrityError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from django.utils import timezone
from weasyprint import HTML
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
import tempfile
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET, require_POST
from django.http import FileResponse, HttpResponseNotFound
from django.conf import settings
from pathlib import Path
from datetime import datetime, time, timedelta
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import login
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.contrib.auth.tokens import default_token_generator
from django.urls import reverse
import json
from django.views.decorators.csrf import csrf_exempt
import cloudinary.uploader
from django.core.files.base import ContentFile
from django.utils.dateparse import parse_date

# ----------------------------
# Login (pantalla)
# ----------------------------
def login_view(request):
    return render(request, "login/login.html")

# ----------------------------
# Iniciar sesión
# ----------------------------
def iniciar_sesion(request):
    if request.method == 'POST':
        correo = request.POST.get('correo', '').strip().lower()
        password = request.POST.get('password', '')

        # BUSCAR USUARIO POR CORREO
        try:
            usuario_db = Usuario.objects.get(correo=correo)
        except Usuario.DoesNotExist:
            usuario_db = None

        # ===============================
        # VALIDACIÓN DE CUENTA INACTIVA
        # ===============================
        if usuario_db and usuario_db.rol == 'cliente' and not usuario_db.is_active:

            # 🔹 Caso 1: cliente se registró pero NO activó el correo
            if usuario_db.cambio_password:
                messages.error(
                    request,
                    "Debes activar tu cuenta desde el correo antes de iniciar sesión."
                )

            # 🔹 Caso 2: cliente desactivó su cuenta voluntariamente
            else:
                messages.error(
                    request,
                    "Tu cuenta está desactivada. Puedes reactivarla desde el correo."
                )

            return redirect('login')

        # ===============================
        # AUTENTICACIÓN NORMAL
        # ===============================
        usuario = authenticate(request, correo=correo, password=password)

        if usuario is not None:
            auth_login(request, usuario)

            # Guardar datos en sesión
            request.session['usuario_id'] = usuario.id
            request.session['usuario_rol'] = usuario.rol
            request.session['usuario_nombre'] = f"{usuario.nombre} {usuario.apellido}"

            # Primer ingreso
            if not usuario.cambio_password:
                return redirect('cambiar_password_primera_vez')

            # Redirección por rol
            if usuario.rol == 'admin':
                return redirect('dashboard_admin')
            elif usuario.rol == 'mesero':
                return redirect('dashboard_mesero')
            elif usuario.rol == 'cocinero':
                return redirect('vista_cocina')
            elif usuario.rol == 'cajero':
                return redirect('dashboard_cajero')
            elif usuario.rol == 'cliente':
                return redirect('dashboard_cliente')
            else:
                messages.error(request, "Rol desconocido.")

        else:
            messages.error(request, "Correo o contraseña incorrectos.")

    return render(request, 'login/login.html')

# ----------------------------
# Cerrar sesión
# ----------------------------
@login_required(login_url='login')
def cerrar_sesion(request):
    auth_logout(request)
    request.session.flush()
    return redirect('login')

# ----------------------------
# Registro del cliente
# ----------------------------

def registro_cliente(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        correo = request.POST.get('correo', '').strip().lower()
        telefono = request.POST.get('telefono', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')

        # ===== VALIDACIONES (LAS TUYAS) =====
        if Usuario.objects.filter(correo=correo).exists():
            messages.error(request, "Ya existe una cuenta con este correo.")
            return redirect('registro_cliente')

        if password != password2:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect('registro_cliente')

        # ===== CREAR USUARIO INACTIVO =====
        usuario = Usuario.objects.create(
            username=correo,
            correo=correo,
            nombre=nombre,
            apellido=apellido,
            telefono=telefono,
            rol='cliente',
            password=make_password(password),
            is_active=False,         
            cambio_password=True
        )

        # ===== TOKEN + UID =====
        uid = urlsafe_base64_encode(force_bytes(usuario.pk))
        token = default_token_generator.make_token(usuario)

        enlace = request.build_absolute_uri(
            reverse('activar_cuenta_cliente', args=[uid, token])
        )

        # ===== ENVIAR CORREO =====
        send_mail(
            subject="Activa tu cuenta - Café Restaurante",
            message=(
                f"Hola {usuario.nombre},\n\n"
                f"Gracias por registrarte.\n\n"
                f"Para activar tu cuenta haz clic en el siguiente enlace:\n\n"
                f"{enlace}\n\n"
                f"Si no te registraste, ignora este correo."
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[correo],
            fail_silently=False,
        )

        messages.success(
            request,
            "Te enviamos un correo para activar tu cuenta."
        )
        return redirect('login')

    return render(request, 'login/registro_cliente.html')

def activar_cuenta_cliente(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(pk=uid, rol='cliente')
    except Exception:
        usuario = None

    if usuario is None or not default_token_generator.check_token(usuario, token):
        messages.error(request, "El enlace de activación no es válido.")
        return redirect('login')

    usuario.is_active = True
    usuario.save(update_fields=['is_active'])

    messages.success(
        request,
        "Cuenta activada correctamente. Ya puedes iniciar sesión."
    )
    return redirect('login')


def verificar_correo_ajax(request):
    correo = request.GET.get('correo', '').strip().lower()

    existe = Usuario.objects.filter(correo=correo).exists()

    return JsonResponse({
        'existe': existe
    })

def cliente_recuperar_password(request):
    if request.method == 'POST':
        correo = request.POST.get('correo', '').strip().lower()

        try:
            usuario = Usuario.objects.get(
                correo=correo,
                rol='cliente',
                is_active=True
            )
        except Usuario.DoesNotExist:
            messages.info(
                request,
                "Si el correo está registrado, recibirás un enlace para restablecer tu contraseña."
            )
            return redirect('login')

        uid = urlsafe_base64_encode(force_bytes(usuario.pk))
        token = default_token_generator.make_token(usuario)

        enlace = request.build_absolute_uri(
            reverse(
                'cliente_restablecer_password',
                args=[uid, token]
            )
        )

        send_mail(
            subject="Recuperación de contraseña - Café Restaurante",
            message=(
                f"Hola {usuario.nombre},\n\n"
                f"Solicitaste restablecer tu contraseña.\n\n"
                f"Haz clic en el siguiente enlace:\n\n"
                f"{enlace}\n\n"
                f"Si no fuiste tú, ignora este mensaje."
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[correo],
            fail_silently=False,
        )

        messages.success(
            request,
            "Te enviamos un enlace a tu correo para restablecer tu contraseña."
        )
        return redirect('login')

    return render(request, 'login/cliente_recuperar_password.html')


def cliente_restablecer_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(pk=uid, rol='cliente')
    except Exception:
        usuario = None

    if usuario is None or not default_token_generator.check_token(usuario, token):
        messages.error(
            request,
            "El enlace es inválido o ya fue utilizado."
        )
        return redirect('login')

    if request.method == 'POST':
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')

        if len(password) < 6:
            messages.error(
                request,
                "La contraseña debe tener al menos 6 caracteres."
            )
            return redirect(request.path)

        if password != password2:
            messages.error(
                request,
                "Las contraseñas no coinciden."
            )
            return redirect(request.path)

        usuario.set_password(password)
        usuario.save(update_fields=['password'])

        messages.success(
            request,
            "Contraseña actualizada correctamente. Ya puedes iniciar sesión."
        )
        return redirect('login')

    return render(
        request,
        'login/cliente_restablecer_password.html'
    )


# ----------------------------
# Cambio de contraseña
# ----------------------------
@login_required(login_url='login')
def cambiar_password_primera_vez(request):
    if request.method == 'POST':
        nueva = request.POST.get('nueva')
        confirmar = request.POST.get('confirmar')

        if nueva and confirmar and nueva == confirmar:
            request.user.set_password(nueva)
            request.user.cambio_password = True
            request.user.save()
            messages.success(request, "Contraseña cambiada correctamente. Inicia sesión nuevamente.")
            return redirect('login')
        else:
            messages.error(request, "Las contraseñas no coinciden.")

    return render(request, 'login/cambiar_password_primera_vez.html')

# ----------------------------
# Dashboards protegidos por rol
# ----------------------------
@login_required(login_url='login')
def dashboard_admin(request):
    if request.user.rol != 'admin':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('login')
    return render(request, 'administrador/dashboard.html')

@login_required(login_url='login')
def dashboard_mesero(request):
    if request.user.rol != 'mesero':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('login')

    return render(request, 'mesero/dashboard.html', {
        "VAPID_PUBLIC_KEY": settings.VAPID_PUBLIC_KEY
    })

@login_required(login_url='login')
def dashboard_cocinero(request):
    if request.user.rol != 'cocinero':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('login')
    return render(request, 'cocinero/dashboard.html')

@login_required(login_url='login')
def dashboard_cajero(request):
    if request.user.rol != 'cajero':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('login')

    return render(request, 'cajero/dashboard.html', {
        "VAPID_PUBLIC_KEY": settings.VAPID_PUBLIC_KEY
    })

@login_required(login_url='login')
def dashboard_cliente(request):
    if request.user.rol != 'cliente':
        messages.error(request, "No tienes permisos para acceder a esta página.")
        return redirect('login')

    return render(request, 'cliente/dashboard.html', {
        "VAPID_PUBLIC_KEY": settings.VAPID_PUBLIC_KEY
    })

# ----------------------------
# Admin-Perfil
# ----------------------------
@login_required(login_url='login')
@rol_requerido('admin')
def perfil_admin(request):
    usuario = request.user   # usuario logueado
    return render(request, "administrador/perfil.html", {"usuario": usuario})

@login_required(login_url='login')
@rol_requerido('admin')
def editar_perfil_admin(request):
    usuario = request.user  # administrador logueado

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        password = request.POST.get("password", "")
        password2 = request.POST.get("password2", "")

        # -----------------------------
        # VALIDACIONES BACKEND
        # -----------------------------

        # Nombre y apellido: solo letras
        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$', nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$', apellido):
            messages.error(request, "El apellido solo debe contener letras.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        # Correo obligatorio + válido
        if not correo:
            messages.error(request, "El correo es obligatorio.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        if not re.match(r"^[^@]+@[^@]+\.[^@]+$", correo):
            messages.error(request, "El formato del correo no es válido.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        # Validar correo no repetido (excepto el mismo usuario)
        if Usuario.objects.exclude(id=usuario.id).filter(correo=correo).exists():
            messages.error(request, "Ya existe un usuario con ese correo.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        # -----------------------------
        # Teléfono obligatorio + validación
        # -----------------------------
        if not telefono:
            messages.error(request, "El teléfono es obligatorio.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        if not telefono.isdigit():
            messages.error(request, "El teléfono solo debe contener números.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        if len(telefono) != 10:
            messages.error(request, "El teléfono debe tener 10 dígitos.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        # -----------------------------
        # Dirección obligatoria
        # -----------------------------
        if not direccion:
            messages.error(request, "La dirección es obligatoria.")
            return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

        # -----------------------------
        # Contraseñas (opcionales)
        # -----------------------------
        if password or password2:
            if len(password) < 6:
                messages.error(request, "La contraseña debe tener mínimo 6 caracteres.")
                return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

            usuario.password = make_password(password)

        # -----------------------------
        # Guardar cambios
        # -----------------------------
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.correo = correo
        usuario.username = correo
        usuario.telefono = telefono
        usuario.direccion = direccion
        usuario.save()

        messages.success(request, "Perfil actualizado correctamente.")
        return redirect("perfil_admin")

    return render(request, "administrador/editar_perfil_admin.html", {"usuario": usuario})

# ----------------------------
# Admin-Gestion de trabajadores
# ----------------------------

@login_required(login_url='login')
@rol_requerido('admin')
def listar_trabajadores(request):
    trabajadores = Usuario.objects.exclude(
        rol__in=['cliente', 'admin']
    ).order_by('id')

    return render(request, 'administrador/trabajadores/trabajadores.html', {
        'trabajadores': trabajadores
    })

@login_required(login_url='login')
@rol_requerido('admin')
def crear_trabajador(request):
    TOPES = {'mesero': 4, 'cajero': 3, 'cocinero': 4}
    horarios = Horario.objects.filter(activo=True)  # Traemos solo horarios activos

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        correo = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        rol = request.POST.get('rol', '').strip()
        horario_id = request.POST.get('horario', '').strip()  # Recibimos el ID del horario
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')

        # Validar campos obligatorios
        if not all([nombre, apellido, correo, telefono, direccion, rol, horario_id, password, password2]):
            messages.error(request, "Todos los campos son obligatorios.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar rol permitido
        if rol not in ['mesero', 'cocinero', 'cajero']:
            messages.error(request, "Rol inválido.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar horario
        try:
            horario_obj = Horario.objects.get(id=horario_id, activo=True)
        except Horario.DoesNotExist:
            messages.error(request, "Seleccione un horario válido.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar tope por rol
        tope = TOPES.get(rol)
        if tope is not None:
            actuales = Usuario.objects.filter(rol=rol, is_active=True).exclude(rol__in=['admin','cliente']).count()
            if actuales >= tope:
                messages.error(request, f"No se puede registrar más personal para el rol {rol}. Tope permitido: {tope}.")
                return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar solo letras
        solo_letras = r'^[a-zA-ZÁÉÍÓÚáéíóúÑñ\s]+$'
        if not re.match(solo_letras, nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        if not re.match(solo_letras, apellido):
            messages.error(request, "El apellido solo debe contener letras.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar email
        try:
            validate_email(correo)
        except ValidationError:
            messages.error(request, "Ingrese un correo electrónico válido.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        if Usuario.objects.filter(correo=correo).exists():
            messages.error(request, "Ya existe un usuario con este correo.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar teléfono
        if not telefono.isdigit() or len(telefono) != 10:
            messages.error(request, "El teléfono debe tener exactamente 10 números.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Validar contraseñas
        if len(password) < 6:
            messages.error(request, "La contraseña debe tener al menos 6 caracteres.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        if password != password2:
            messages.error(request, "Las contraseñas no coinciden.")
            return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})

        # Crear trabajador
        trabajador = Usuario(
            nombre=nombre,
            apellido=apellido,
            correo=correo,
            username=correo,
            telefono=telefono,
            direccion=direccion,
            rol=rol,
            horario=horario_obj
        )
        trabajador.set_password(password)
        trabajador.save()

        messages.success(request, f"Trabajador {nombre} {apellido} creado con éxito.")
        return redirect('listar_trabajadores')

    return render(request, 'administrador/trabajadores/crear_trabajador.html', {'horarios': horarios})


@login_required(login_url='login')
@rol_requerido('admin')
def editar_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Usuario, id=trabajador_id)
    TOPES = {'mesero': 4, 'cajero': 3, 'cocinero': 4}
    horarios = Horario.objects.filter(activo=True)  # Solo horarios activos

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        correo = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        rol = request.POST.get('rol', '').strip()
        horario_id = request.POST.get('horario', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')

        # Validar rol
        if rol not in ['mesero', 'cocinero', 'cajero']:
            messages.error(request, "Rol inválido.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Validar horario
        try:
            horario_obj = Horario.objects.get(id=horario_id, activo=True)
        except Horario.DoesNotExist:
            messages.error(request, "Seleccione un horario válido.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Validar tope por rol si cambia
        if rol != trabajador.rol:
            tope = TOPES.get(rol)
            if tope is not None:
                actuales = Usuario.objects.filter(rol=rol, is_active=True).exclude(id=trabajador.id).count()
                if actuales >= tope:
                    messages.error(request, f"No se puede asignar el rol {rol}. Ya se alcanzó el tope: {tope}.")
                    return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Validar solo letras
        solo_letras = r'^[a-zA-ZÁÉÍÓÚáéíóúÑñ\s]+$'
        if not re.match(solo_letras, nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        if not re.match(solo_letras, apellido):
            messages.error(request, "El apellido solo debe contener letras.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Validar email
        try:
            validate_email(correo)
        except ValidationError:
            messages.error(request, "Ingrese un correo electrónico válido.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        if Usuario.objects.exclude(id=trabajador.id).filter(correo=correo).exists():
            messages.error(request, "Ya existe un usuario con este correo.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Validar teléfono
        if not telefono.isdigit() or len(telefono) != 10:
            messages.error(request, "El teléfono debe tener exactamente 10 números.")
            return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

        # Contraseñas si se envían
        if password or password2:
            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

            if len(password) < 6:
                messages.error(request, "La contraseña debe tener al menos 6 caracteres.")
                return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

            trabajador.set_password(password)

        # Actualizar trabajador
        trabajador.nombre = nombre
        trabajador.apellido = apellido
        trabajador.correo = correo
        trabajador.telefono = telefono
        trabajador.direccion = direccion
        trabajador.rol = rol
        trabajador.horario = horario_obj
        trabajador.save()

        messages.success(request, f"Trabajador {trabajador.nombre} actualizado con éxito.")
        return redirect('listar_trabajadores')

    return render(request, 'administrador/trabajadores/editar_trabajador.html', {'trabajador': trabajador, 'horarios': horarios})

@login_required(login_url='login')
@rol_requerido('admin')
def eliminar_trabajador(request, trabajador_id):
    """
    Si el trabajador nunca ha interactuado en el sistema, se elimina.
    Si ya tiene interacciones (pedidos, mensajes, notificaciones),
    solo se desactiva (is_active = False).
    """
    if request.method == 'POST':
        trabajador = get_object_or_404(Usuario, id=trabajador_id)
        nombre_trabajador = f"{trabajador.nombre} {trabajador.apellido}"

        # Solo aplicar a roles de trabajador
        if trabajador.rol not in ['mesero', 'cocinero', 'cajero']:
            messages.error(request, "No se puede eliminar este usuario.")
            return redirect('listar_trabajadores')

        # Usamos la propiedad del modelo
        tiene_interacciones = trabajador.tiene_interacciones

        if tiene_interacciones:
            # NO se elimina, solo se desactiva
            if trabajador.is_active:
                trabajador.is_active = False
                trabajador.save()
                messages.warning(
                    request,
                    f"El trabajador {nombre_trabajador} tiene registros asociados, "
                    "por lo que no se eliminó. Se desactivó su cuenta."
                )
            else:
                messages.info(
                    request,
                    f"El trabajador {nombre_trabajador} ya estaba desactivado."
                )
        else:
            # Sin interacciones → se puede eliminar físicamente
            trabajador.delete()
            messages.success(
                request,
                f"Trabajador {nombre_trabajador} eliminado correctamente."
            )

    return redirect('listar_trabajadores')

@login_required(login_url='login')
@rol_requerido('admin')
def activar_trabajador(request, trabajador_id):
    if request.method != 'POST':
        return redirect('listar_trabajadores')

    TOPES = {
        'mesero': 4,
        'cajero': 3,
        'cocinero': 4,
    }

    # Bloqueo transaccional para evitar activar 2 al mismo tiempo y pasarse del tope
    with transaction.atomic():
        trabajador = get_object_or_404(
            Usuario.objects.select_for_update(),
            id=trabajador_id
        )

        # Solo aplicar a roles de trabajador
        if trabajador.rol not in ['mesero', 'cocinero', 'cajero']:
            messages.error(request, "No se puede activar este usuario.")
            return redirect('listar_trabajadores')

        # Si ya está activo, no hacer nada
        if trabajador.is_active:
            messages.info(
                request,
                f"El trabajador {trabajador.nombre} {trabajador.apellido} ya estaba activo."
            )
            return redirect('listar_trabajadores')

        # ===== VALIDAR TOPE POR ROL =====
        tope = TOPES.get(trabajador.rol)
        if tope is not None:
            actuales = (
                Usuario.objects
                .filter(rol=trabajador.rol, is_active=True)
                .exclude(id=trabajador.id)
                .count()
            )

            if actuales >= tope:
                messages.error(
                    request,
                    f"No se puede activar a {trabajador.nombre} {trabajador.apellido}. "
                    f"Ya se alcanzó el tope para el rol {trabajador.rol} (máximo {tope})."
                )
                return redirect('listar_trabajadores')

        # Activar
        trabajador.is_active = True
        trabajador.save()

    messages.success(
        request,
        f"Trabajador {trabajador.nombre} {trabajador.apellido} activado correctamente."
    )
    return redirect('listar_trabajadores')

# ----------------------------
# Admin - Gestión de horario
# ----------------------------

DIAS_VALIDOS = [
    "Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo", "lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo",
]

# ----------------------------
# Listar horarios
# ----------------------------
def lista_horarios(request):
    horarios = Horario.objects.all().order_by('id')
    return render(request, 'administrador/horario/horario.html', {'horarios': horarios})

# ----------------------------
# Crear horario
# ----------------------------
def crear_horario(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')
        dias = request.POST.get('dias')
        activo = True  # siempre activo al crear

        # Validar que hora fin >= hora inicio
        fmt = "%H:%M"
        try:
            hi = datetime.strptime(hora_inicio, fmt).time()
            hf = datetime.strptime(hora_fin, fmt).time()
            if hf <= hi:
                messages.error(request, "La hora fin no puede ser menor o igual a la hora inicio.")
                return render(request, 'administrador/horario/crear_horario.html')
        except ValueError:
            messages.error(request, "Formato de hora inválido.")
            return render(request, 'administrador/horario/crear_horario.html')

        # Validar que las horas correspondan al horario seleccionado
        if not validar_horario_por_nombre(nombre, hora_inicio, hora_fin):
            messages.error(request, f"El horario seleccionado {nombre} no coincide con las horas ingresadas.")
            return render(request, 'administrador/horario/crear_horario.html')

        # Validar días
        if not validar_dias(dias):
            messages.error(request, f"Los días ingresados no son válidos. Deben ser nombres correctos de los días.")
            return render(request, 'administrador/horario/crear_horario.html')

        Horario.objects.create(
            nombre=nombre,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            dias=dias,
            activo=activo
        )
        messages.success(request, f"Horario {nombre} creado correctamente.")
        return redirect('lista_horarios')
    
    return render(request, 'administrador/horario/crear_horario.html')

# ----------------------------
# Editar horario
# ----------------------------
def editar_horario(request, id):
    horario = get_object_or_404(Horario, id=id)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')
        dias = request.POST.get('dias')

        # Validar que hora fin >= hora inicio
        fmt = "%H:%M"
        try:
            hi = datetime.strptime(hora_inicio, fmt).time()
            hf = datetime.strptime(hora_fin, fmt).time()
            if hf <= hi:
                messages.error(request, "La hora fin no puede ser menor o igual a la hora inicio.")
                return render(request, 'administrador/horario/editar_horario.html', {'horario': horario})
        except ValueError:
            messages.error(request, "Formato de hora inválido.")
            return render(request, 'administrador/horario/editar_horario.html', {'horario': horario})

        # Validar que las horas correspondan al horario seleccionado
        if not validar_horario_por_nombre(nombre, hora_inicio, hora_fin):
            messages.error(request, f"El horario seleccionado {nombre} no coincide con las horas ingresadas.")
            return render(request, 'administrador/horario/editar_horario.html', {'horario': horario})

        # Validar días
        if not validar_dias(dias):
            messages.error(request, f"Los días ingresados no son válidos. Deben ser nombres correctos de los días.")
            return render(request, 'administrador/horario/editar_horario.html', {'horario': horario})

        horario.nombre = nombre
        horario.hora_inicio = hora_inicio
        horario.hora_fin = hora_fin
        horario.dias = dias

        # Activo siempre True mientras tenga trabajadores asignados
        horario.activo = True

        horario.save()
        messages.success(request, f"Horario {nombre} actualizado correctamente.")
        return redirect('lista_horarios')

    return render(request, 'administrador/horario/editar_horario.html', {'horario': horario})

# ----------------------------
# Eliminar horario
# ----------------------------
def eliminar_horario(request, id):
    horario = get_object_or_404(Horario, id=id)

    if horario.tiene_trabajadores_asignados():
        messages.error(
            request,
            f"No se puede eliminar el horario {horario.nombre} porque tiene trabajadores asignados. "
            "Solo se permite editar."
        )
        return redirect('lista_horarios')

    # Si no tiene trabajadores → sí se puede eliminar
    horario.delete()
    messages.success(request, f"Horario {horario.nombre} eliminado correctamente.")
    return redirect('lista_horarios')

# ----------------------------
# Función auxiliar para validar horario por nombre
# ----------------------------
def validar_horario_por_nombre(nombre, hora_inicio, hora_fin):
    """
    Retorna True si las horas corresponden al tipo de horario:
    Mañana (AM), Tarde (PM temprano), Noche (PM tarde)
    """
    fmt = "%H:%M"
    try:
        hi = datetime.strptime(hora_inicio, fmt).time()
        hf = datetime.strptime(hora_fin, fmt).time()
    except ValueError:
        return False

    if nombre == "Mañana":
        return hi >= datetime.strptime("06:00", fmt).time() and hf <= datetime.strptime("12:00", fmt).time()
    elif nombre == "Tarde":
        return hi >= datetime.strptime("12:00", fmt).time() and hf <= datetime.strptime("18:00", fmt).time()
    elif nombre == "Noche":
        return hi >= datetime.strptime("18:00", fmt).time() and hf <= datetime.strptime("23:59", fmt).time()
    else:
        return False

# ----------------------------
# Función auxiliar para validar días
# ----------------------------
def validar_dias(dias_texto):
    """
    Acepta:
    - Lunes, Martes, Miercoles
    - Lunes y Martes
    - Lunes, Martes y Miercoles
    - Lunes a Viernes
    - Lunes a Viernes, Sabado y Domingo
    """
    texto = dias_texto.strip().lower()

    # Normalizar: cambiar " y " por coma
    texto = texto.replace(" y ", ",")

    partes = [p.strip() for p in texto.split(",")]

    for parte in partes:
        if not parte:
            continue

        # Rango: solo si tiene " a "
        if " a " in parte:
            inicio, fin = [p.strip().capitalize() for p in parte.split(" a ")]
            if inicio not in DIAS_VALIDOS or fin not in DIAS_VALIDOS:
                return False
        else:
            dia = parte.capitalize()
            if dia not in DIAS_VALIDOS:
                return False

    return True

# ----------------------------
# Admin - Gestión de Mesas
# ----------------------------
# Listar mesas
@login_required(login_url='login')
@rol_requerido('admin')
def listar_mesas(request):
    mesas = Mesa.objects.all().order_by('numero')
    return render(request, 'administrador/mesas/listar_mesas.html', {'mesas': mesas})

#crear una mesa
@login_required(login_url='login')
@rol_requerido('admin')
def registrar_mesa(request):
    # Obtener el último número de mesa
    ultima_mesa = Mesa.objects.order_by('-numero').first()
    siguiente_numero = ultima_mesa.numero + 1 if ultima_mesa else 1

    if request.method == 'POST':
        numero = request.POST.get('numero')
        capacidad = request.POST.get('capacidad')

        # Validar campos vacíos
        if not numero or not capacidad:
            messages.error(request, "Todos los campos son obligatorios.")
            return render(request, 'administrador/mesas/registrar_mesa.html', {
                'siguiente_numero': siguiente_numero
            })

        # Validar que sean números enteros
        if not numero.isdigit() or not capacidad.isdigit():
            messages.error(request, "El número y la capacidad deben ser valores numéricos.")
            return render(request, 'administrador/mesas/registrar_mesa.html', {
                'siguiente_numero': siguiente_numero
            })

        # Convertir a int para validaciones
        numero = int(numero)
        capacidad = int(capacidad)

        # Validar que sean positivos
        if numero < 1 or capacidad < 1:
            messages.error(request, "Número y capacidad deben ser mayores o iguales a 1.")
            return render(request, 'administrador/mesas/registrar_mesa.html', {
                'siguiente_numero': siguiente_numero
            })

        # Validar número único
        if Mesa.objects.filter(numero=numero).exists():
            messages.error(request, "Ya existe una mesa con ese número.")
            return render(request, 'administrador/mesas/registrar_mesa.html', {
                'siguiente_numero': siguiente_numero
            })

        # Crear mesa
        Mesa.objects.create(
            numero=numero,
            capacidad=capacidad,
        )
        messages.success(request, f"Mesa {numero} registrada correctamente.")
        return redirect('listar_mesas')

    return render(request, 'administrador/mesas/registrar_mesa.html', {
        'siguiente_numero': siguiente_numero
    })

# Editar mesa
@login_required(login_url='login')
@rol_requerido('admin')
def editar_mesa(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)

    if request.method == 'POST':
        numero = request.POST.get('numero')
        capacidad = request.POST.get('capacidad')

        # Evitar que alteren el número
        if str(mesa.numero) != str(numero):
            messages.error(request, "No es posible modificar el número de mesa.")
            return render(request, 'administrador/mesas/editar_mesa.html', {
                'mesa': mesa
            })

        # Validar campos vacíos
        if not capacidad:
            messages.error(request, "La capacidad es obligatoria.")
            return render(request, 'administrador/mesas/editar_mesa.html', {
                'mesa': mesa
            })

        # Validar numérico
        if not capacidad.isdigit():
            messages.error(request, "La capacidad debe ser un número válido.")
            return render(request, 'administrador/mesas/editar_mesa.html', {
                'mesa': mesa
            })

        capacidad = int(capacidad)

        # Validar mínimo
        if capacidad < 1:
            messages.error(request, "La capacidad debe ser mínimo 1.")
            return render(request, 'administrador/mesas/editar_mesa.html', {
                'mesa': mesa
            })

        mesa.capacidad = capacidad
        mesa.save()

        messages.success(request, f"Mesa {mesa.numero} actualizada correctamente.")
        return redirect('listar_mesas')

    return render(request, 'administrador/mesas/editar_mesa.html', {
        'mesa': mesa
    })

# Eliminar mesa
@login_required(login_url='login')
@rol_requerido('admin')
def eliminar_mesa(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)

    if request.method == 'POST':
        # ====== VALIDAR SI LA MESA ESTÁ OCUPADA ======
        # pedidos que aún no han terminado
        pedidos_activos = Pedido.objects.filter(
            mesa=mesa
        ).exclude(
            estado__in=['finalizado', 'cancelado']  # <-- ajusta a tus estados reales
        ).exists()

        if pedidos_activos:
            messages.error(
                request,
                f"No puedes eliminar la mesa {mesa.numero} porque tiene pedidos activos (mesa ocupada)."
            )
            return redirect('listar_mesas')

        # Si no tiene pedidos activos, se puede eliminar
        numero = mesa.numero
        mesa.delete()
        messages.success(request, f"Mesa {numero} eliminada correctamente.")
        return redirect('listar_mesas')

    # Confirmación opcional
    return render(request, 'administrador/mesas/confirmar_eliminar.html', {'mesa': mesa})

# ----------------------------
# Admin - Gestión de Menú / Productos
# ----------------------------
# Listar productos
@login_required(login_url='login')
@rol_requerido('admin')
def listar_menu(request):
    productos = Producto.objects.all().order_by('id')
    tipo = request.GET.get("tipo")
    if tipo:
        productos = productos.filter(tipo=tipo)
    return render(request, 'administrador/menu/listar_menu.html', {'productos': productos})

# Registrar producto
@login_required(login_url='login')
@rol_requerido('admin')
def registrar_menu(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        precio = request.POST.get('precio', '').strip()
        tipo = request.POST.get('tipo', '').strip()

        # 1. VALIDAR CAMPOS VACÍOS PRIMERO
        if not nombre or not precio or not tipo:
            messages.error(request, "Todos los campos son obligatorios.")
            return render(request, 'administrador/menu/registrar_menu.html', {
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        # 2. VALIDAR SOLO LETRAS
        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ ]+$', nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, 'administrador/menu/registrar_menu.html', {
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        # 3. VALIDAR PRECIO
        try:
            precio = float(precio)
            if precio < 0.50 or precio > 15.99:
                raise ValueError
        except ValueError:
            messages.error(request, "El precio debe ser un número válido entre 0.50 y 15.99.")
            return render(request, 'administrador/menu/registrar_menu.html', {
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': request.POST.get('precio', ''),
                'tipo': tipo,
            })


        # 4. VALIDAR NOMBRE DUPLICADO
        if Producto.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, "Ya existe un producto con ese nombre.")
            return render(request, 'administrador/menu/registrar_menu.html', {
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        # 5. CREAR PRODUCTO
        Producto.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            precio=precio,
            tipo=tipo,
        )

        messages.success(request, f"Producto {nombre} registrado correctamente.")
        return redirect('listar_menu')

    return render(request, 'administrador/menu/registrar_menu.html')

# Editar producto
@login_required(login_url='login')
@rol_requerido('admin')
def editar_menu(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        precio = request.POST.get('precio', '').strip()
        tipo = request.POST.get('tipo', '').strip()

        # -------- VALIDACIONES BACKEND --------

        # Validar campos obligatorios
        if not nombre or not precio or not tipo:
            messages.error(request, "Todos los campos obligatorios deben completarse.")
            return render(request, 'administrador/menu/editar_menu.html', {
                'producto': producto,
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        # Validar que el nombre contenga solo letras
        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ ]+$', nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, 'administrador/menu/editar_menu.html', {
                'producto': producto,
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        # Validar precio
        try:
            precio_float = float(precio)
            if precio < 0.50 or precio > 15.99:
                raise ValueError
        except ValueError:
            messages.error(request, "El precio debe ser un número válido entre 0.50 y 15.99.")
            return render(request, 'administrador/menu/editar_menu.html', {
                'producto': producto,
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })


        # Validar nombre repetido excepto el item actual
        if Producto.objects.exclude(id=producto.id).filter(nombre__iexact=nombre).exists():
            messages.error(request, "Ya existe otro producto con ese nombre.")
            return render(request, 'administrador/menu/editar_menu.html', {
                'producto': producto,
                'nombre': nombre,
                'descripcion': descripcion,
                'precio': precio,
                'tipo': tipo,
            })

        producto.nombre = nombre
        producto.descripcion = descripcion
        producto.precio = precio
        producto.tipo = tipo
        producto.save()

        messages.success(request, f"Producto {nombre} actualizado correctamente.")
        return redirect('listar_menu')

    return render(request, 'administrador/menu/editar_menu.html', {'producto': producto})

# Eliminar producto
@login_required(login_url='login')
@rol_requerido('admin')
def eliminar_menu(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    # Solo debería llegar por POST desde el formulario con SweetAlert
    if request.method != 'POST':
        return redirect('listar_menu')

    # Verificar si el producto tiene detalles de pedido asociados
    tiene_pedidos = producto.tiene_pedidos

    if tiene_pedidos:
        # No eliminar, solo desactivar
        if producto.activo:
            producto.activo = False
            producto.save()
            messages.warning(
                request,
                f"El producto {producto.nombre} tiene pedidos asociados, "
                "por lo que no se eliminó. Se desactivó del menú."
            )
        else:
            messages.info(
                request,
                f"El producto {producto.nombre} ya estaba desactivado."
            )
    else:
        nombre = producto.nombre
        producto.delete()
        messages.success(
            request,
            f"Producto {nombre} eliminado correctamente."
        )

    return redirect('listar_menu')

# Activar producto
@login_required(login_url='login')
@rol_requerido('admin')
def activar_menu(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method != 'POST':
        return redirect('listar_menu')

    if not producto.activo:
        producto.activo = True
        producto.save()
        messages.success(request, f"Producto {producto.nombre} activado correctamente.")
    else:
        messages.info(request, f"El producto {producto.nombre} ya estaba activo.")

    return redirect('listar_menu')

#-----------------------------
# Admin-Clientes
#-----------------------------

@login_required(login_url='login')
@rol_requerido('admin')
def admin_listar_clientes(request):

    buscar = request.GET.get('buscar', '').strip()

    clientes = Usuario.objects.filter(rol='cliente')

    if buscar:
        clientes = clientes.filter(
            nombre__icontains=buscar
        ) | clientes.filter(
            apellido__icontains=buscar
        ) | clientes.filter(
            correo__icontains=buscar
        )

    clientes = clientes.order_by('nombre', 'apellido')

    return render(request, 'administrador/clientes/listar.html', {
        'clientes': clientes,
        'buscar': buscar
    })

# ----------------------------
# Mesero-Pedidos
# ----------------------------
@login_required(login_url='login')
@rol_requerido('mesero')
def api_estados_pedidos(request):
    hoy = timezone.localdate()

    pedidos = Pedido.objects.filter(
        tipo_pedido='restaurante',
        mesero=request.user,
        fecha_hora__date=hoy
    ).select_related("mesa").values(
        "id",
        "estado",
        "mesa_id",
        "mesa__estado",
    )

    return JsonResponse(list(pedidos), safe=False)

# Listar pedidos
@login_required(login_url='login')
@rol_requerido('mesero')
def listar_pedidos(request):

    hoy = timezone.localdate()  # Fecha del sistema
    pedidos = Pedido.objects.filter(
        tipo_pedido='restaurante',
        mesero=request.user,
        fecha_hora__date=hoy      
    ).order_by('id')

    return render(request, 'mesero/pedidos/listar_pedidos.html', {
        'pedidos': pedidos
    })

# Crear pedido
@login_required(login_url='login')
@rol_requerido('mesero')
def crear_pedido(request):
    if request.method == 'POST':
        mesero_id = request.user.id
        mesa_id = request.POST.get('mesa')

        if not mesa_id:
            messages.error(request, "Debe seleccionar una mesa.")
            return redirect('crear_pedido')

        # Crear el pedido
        pedido = Pedido.objects.create(
            mesero_id=mesero_id,
            mesa_id=mesa_id,
            tipo_pedido="restaurante",
            estado="en_creacion"
        )

        #CAMBIAR MESA A OCUPADA
        mesa = Mesa.objects.get(id=mesa_id)
        mesa.estado = "ocupada"
        mesa.save()

        return redirect('agregar_detalles', pedido_id=pedido.id)

    #SOLO MOSTRAR MESAS LIBRES
    mesas = Mesa.objects.filter(estado="libre")

    return render(request, 'mesero/pedidos/crear_pedido.html', {
        'mesas': mesas,
        'user': request.user
    })

# Ver pedido
@login_required(login_url='login')
@rol_requerido('mesero')
def ver_pedido(request, pedido_id):
    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        mesero=request.user,
        tipo_pedido='restaurante'
    )

    # Validación: no permitir ver pedido sin productos
    if not pedido.detalles.exists():
        messages.error(request, "Debe agregar al menos un producto antes de finalizar.")
        return redirect('agregar_detalles', pedido_id=pedido.id)

    # SOLO la primera vez (cuando aún es en_creacion) se manda a cocina
    if pedido.estado == "en_creacion" and not pedido.enviado_cocina:
        pedido.enviado_cocina = True
        pedido.save(update_fields=["enviado_cocina"])
        enviar_pedido_cocina(pedido)

    # IMPORTANTE: NO mandar actualizar aquí
    # (porque "Ver" se abre muchas veces y eso genera duplicados)

    return render(request, 'mesero/pedidos/ver_pedido.html', {'pedido': pedido})

# Editar pedido
@login_required(login_url='login')
@rol_requerido('mesero')
def editar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        mesero=request.user,
        tipo_pedido="restaurante"
    )

    # SOLO permitir editar si está en creación
    if pedido.estado != "en_creacion":
        return JsonResponse({
            "success": False,
            "mensaje": "Solo se puede editar un pedido cuando está en creación."
        }, status=403)

    if request.method == 'POST':
        nueva_mesa_id = request.POST.get('mesa')
        mesa_anterior_id = pedido.mesa_id

        if not nueva_mesa_id:
            return JsonResponse({"success": False, "mensaje": "Debe seleccionar una mesa."}, status=400)

        # 1) CAMBIO DE MESA → Liberar anterior y ocupar nueva
        if str(mesa_anterior_id) != str(nueva_mesa_id):

            if mesa_anterior_id:
                mesa_old = Mesa.objects.get(id=mesa_anterior_id)
                mesa_old.estado = "libre"
                mesa_old.save(update_fields=["estado"])

            mesa_new = Mesa.objects.get(id=nueva_mesa_id)
            mesa_new.estado = "ocupada"
            mesa_new.save(update_fields=["estado"])

            pedido.mesa_id = nueva_mesa_id

        # 2) NO CAMBIAR ESTADO
        pedido.save(update_fields=["mesa_id"])

        # 3) Enviar actualización a cocina (manteniendo estado en_creacion)
        enviar_actualizacion_cocina(pedido)

        return JsonResponse({
            "success": True,
            "mensaje": f"Pedido #{pedido.id} actualizado correctamente.",
            "pedido_id": pedido.id
        })

    mesas = Mesa.objects.filter(estado="libre") | Mesa.objects.filter(id=pedido.mesa_id)
    productos = Producto.objects.filter(activo=True).order_by("nombre")

    return render(request, 'mesero/pedidos/editar_pedido.html', {
        'pedido': pedido,
        'mesas': mesas,
        'productos': productos
    })


def enviar_actualizacion_cocina(pedido):
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "actualizar_pedido",
            "pedido": {
                "id": pedido.id,
                # Mantener título igual al original
                "mesa": pedido.mesa.numero if pedido.mesa else "Domicilio",

                # Mostrar correctamente quién atendió
                "mesero": pedido.mesero.nombre if pedido.tipo_pedido == 'restaurante' else f"Cajero: {pedido.cajero.nombre}",

                # Productos del pedido
                "productos": [
                    {
                        "nombre": d.producto.nombre,
                        "cantidad": d.cantidad,
                        "observacion": d.observacion,
                    }
                    for d in pedido.detalles.all()
                ],

                # MUY IMPORTANTE: indicar tipo de pedido
                "tipo": pedido.tipo_pedido
            }
        }
    )

# Eliminar pedido
@login_required(login_url='login')
@rol_requerido('mesero')
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        mesero=request.user,
        tipo_pedido="restaurante"
    )


    if pedido.estado != 'en_creacion':
        return JsonResponse({
            "success": False,
            "message": "Solo se puede eliminar un pedido cuando está en creación."
        })

    if request.method == 'POST':

        #LIBERAR LA MESA
        if pedido.mesa:
            pedido.mesa.estado = "libre"
            pedido.mesa.save()

        #ELIMINAR PEDIDO
        pedido_id = pedido.id
        pedido.delete()

        #ENVIAR A COCINA QUE SE ELIMINÓ
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "pedidos_activos",
            {
                "type": "eliminar_pedido",
                "pedido_id": pedido_id
            }
        )

        return JsonResponse({"success": True})

    return JsonResponse({"success": False})

#finalizar un pedido
@login_required(login_url='login')
@rol_requerido('mesero')
def finalizar_pedido(request, pedido_id):

    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Método no permitido."})

    pedido = get_object_or_404(Pedido, id=pedido_id)

    # CAMBIAR ESTADO DEL PEDIDO
    pedido.estado = "finalizado"
    pedido.save()

    # LIBERAR MESA
    if pedido.mesa:
        pedido.mesa.estado = "libre"
        pedido.mesa.save()

    return JsonResponse({
        "success": True,
        "message": f"El pedido #{pedido.id} ha sido finalizado y la mesa está libre."
    })

# ----------------------------
# Detalles con AJAX
# ----------------------------
# Agregar detalles con AJAX
@login_required(login_url='login')
@rol_requerido('mesero')
def agregar_detalles(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # SOLO productos activos
    productos = Producto.objects.filter(activo=True).order_by("nombre")

    return render(request, 'mesero/pedidos/agregar_detalles.html', {
        'pedido': pedido,
        'productos': productos
    })

@login_required(login_url='login')
@rol_requerido('mesero')
def agregar_detalle_ajax(request, pedido_id):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))
        observacion = data.get("observacion", '')

        # VALIDAR: NO permitir agregar si está agotado hoy
        producto = get_object_or_404(Producto, id=producto_id)
        if producto.agotado_hoy:
            return JsonResponse({
                "success": False,
                "mensaje": f'El producto "{producto.nombre}" está AGOTADO hoy y no se puede agregar.'
            })


        # 1) VERIFICAR SI YA EXISTE ESE PRODUCTO EN EL PEDIDO
        if DetallePedido.objects.filter(pedido_id=pedido_id, producto_id=producto_id).exists():
            return JsonResponse({
                'success': False,
                'mensaje': 'Este producto ya fue agregado al pedido. Edítelo en la tabla.'
            })

        # 2) CREAR DETALLE (SIN RECARGOS)
        detalle = DetallePedido.objects.create(
            pedido_id=pedido_id,
            producto_id=producto_id,
            cantidad=cantidad,
            observacion=observacion
        )

        # 3) ACTUALIZAR TOTALES
        pedido = get_object_or_404(Pedido, id=pedido_id)
        pedido.calcular_totales()

        # 4) ORDENAR DETALLES POR ORDEN DE REGISTRO
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('mesero/pedidos/tabla_detalles.html', {
            'pedido': pedido,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Producto agregado.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Error al agregar producto'})

@login_required(login_url='login')
@rol_requerido('mesero')
def editar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        detalle.cantidad = int(data.get("cantidad", detalle.cantidad))
        detalle.observacion = data.get("observacion", detalle.observacion)

        detalle.save()

        # RECALCULAR TOTALES
        pedido.calcular_totales()

        # ORDENAR DETALLES POR ID (ORDEN DE REGISTRO)
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('mesero/pedidos/tabla_detalles.html', {
            'pedido': pedido,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Detalle actualizado correctamente.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Error al actualizar detalle.'})

@login_required(login_url='login')
@rol_requerido('mesero')
def eliminar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if request.method == 'POST':
        detalle.delete()

        # RECALCULAR TOTALES
        pedido.calcular_totales()

        # ORDENAR DETALLES POR ID
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('mesero/pedidos/tabla_detalles.html', {
            'pedido': pedido,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Producto eliminado correctamente.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Error al eliminar producto.'})

@login_required(login_url='login')
@rol_requerido('mesero')
@require_POST
def mesero_cancelar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='restaurante',
        mesero=request.user
    )

    # Si quieres limitar qué estados se pueden cancelar, usa esto:
    # if pedido.estado not in ["en_creacion", "en preparacion", "listo"]:
    #     messages.error(request, "No se puede cancelar este pedido en su estado actual.")
    #     return redirect('listar_pedidos')

    # 1) Liberar mesa
    if pedido.mesa:
        pedido.mesa.estado = "libre"
        pedido.mesa.save(update_fields=["estado"])

    # 2) Guardar id para avisar por WS antes de borrar
    pid = pedido.id

    # 3) Eliminar pedido (borra detalles por cascade)
    pedido.delete()

    # 4) Avisar a cocina en tiempo real para quitarlo de pantalla
    channel_layer = get_channel_layer()
    try:
        async_to_sync(channel_layer.group_send)(
            "pedidos_activos",
            {
                "type": "eliminar_pedido",
                "pedido_id": pid
            }
        )
    except Exception:
        pass

    messages.info(request, "Pedido cancelado correctamente.")
    return redirect('listar_pedidos')

@login_required(login_url='login')
@rol_requerido('mesero')
def mesero_historial_pedidos(request):

    codigo = request.GET.get("codigo", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    metodo = request.GET.get("metodo", "")

    pedidos = (
        Pedido.objects
        .filter(
            tipo_pedido="restaurante",
            mesero=request.user,
            estado="finalizado",
            pagos__estado_pago="confirmado"
        )
        .select_related("mesa")
        .prefetch_related(
            Prefetch(
                "pagos",
                queryset=Pago.objects.filter(estado_pago="confirmado")
                .prefetch_related("comprobante")
            )
        )
        .distinct()
        .order_by("-fecha_hora")
    )

    # ================= FILTROS =================
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo:
        pedidos = pedidos.filter(pagos__metodo_pago=metodo)

    # Pago + comprobante (1 por pedido)
    for p in pedidos:
        p.pago_confirmado = (
            p.pagos.filter(estado_pago="confirmado").order_by("id").first()
        )
        p.comprobante = (
            p.pago_confirmado.comprobante.first()
            if p.pago_confirmado else None
        )

    return render(request, "mesero/pedidos/historial_pedidos.html", {
        "pedidos": pedidos,
        "filtros": {
            "codigo": codigo,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "metodo": metodo,
        }
    })

# ----------------------------
# Cocinero - Perfil
# ----------------------------
@login_required(login_url='login')
@rol_requerido('cocinero')
def perfil_cocinero(request):
    usuario = request.user
    return render(request, "cocinero/perfil/perfil_cocinero.html", {
        "usuario": usuario
    })

@login_required(login_url='login')
@rol_requerido('cocinero')
def editar_perfil_cocinero(request):
    usuario = request.user

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        direccion = request.POST.get("direccion", "").strip()

        password = request.POST.get("password", "").strip()
        password2 = request.POST.get("password2", "").strip()

        # =========================
        # VALIDACIONES BACKEND
        # =========================

        # Nombre y apellido solo letras y espacios (con tildes)
        patron_letras = r"^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$"
        if not re.match(patron_letras, nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

        if not re.match(patron_letras, apellido):
            messages.error(request, "El apellido solo debe contener letras.")
            return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

        # Correo válido
        try:
            validate_email(correo)
        except ValidationError:
            messages.error(request, "Ingrese un correo válido.")
            return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

        # Teléfono: solo dígitos y exactamente 10
        if not telefono.isdigit():
            messages.error(request, "El teléfono solo debe contener números.")
            return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

        if len(telefono) != 10:
            messages.error(request, "El teléfono debe tener exactamente 10 dígitos.")
            return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

        # =========================
        # ASIGNAR CAMPOS
        # =========================
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.correo = correo
        usuario.telefono = telefono
        usuario.direccion = direccion

        # Password opcional
        if password or password2:
            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

            if len(password) < 6:
                messages.error(request, "La contraseña debe tener mínimo 6 caracteres.")
                return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

            usuario.set_password(password)
            usuario.save()
            update_session_auth_hash(request, usuario)
        else:
            usuario.save()

        messages.success(request, "Perfil actualizado correctamente.")
        return redirect("perfil_cocinero")

    return render(request, "cocinero/perfil/editar_perfil_cocinero.html", {"usuario": usuario})

# ----------------------------
# Cocinero marca pedido como listo
# ----------------------------

@login_required(login_url='login')
@rol_requerido('cocinero')
def vista_cocina(request):
    hoy = timezone.localdate()
    estados = ["en_creacion", "aceptado", "en preparacion"]

    pedidos_restaurante = Pedido.objects.filter(
        estado__in=estados,
        tipo_pedido='restaurante',
        fecha_hora__date=hoy
    ).order_by('id')

    pedidos_domicilio = Pedido.objects.filter(
        estado__in=estados,
        tipo_pedido='domicilio',
        fecha_hora__date=hoy
    ).order_by('id')

    meseros = Usuario.objects.filter(rol='mesero')
    cajeros = Usuario.objects.filter(rol='cajero')

    # NUEVO: menú
    productos_menu = Producto.objects.filter(activo=True).order_by('nombre')  # ajusta "activo" si no existe

    return render(request, 'cocinero/pedido.html', {
        'pedidos_restaurante': pedidos_restaurante,
        'pedidos_domicilio': pedidos_domicilio,
        'meseros': meseros,
        'cajeros': cajeros,
        'productos_menu': productos_menu,  # NUEVO
    })

@login_required(login_url='login')
@rol_requerido('cocinero')
@csrf_protect
@require_POST
def marcar_pedido_listo(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado != "en preparacion":
        return JsonResponse({"success": False, "mensaje": "Solo se puede marcar listo si está en preparación."})

    pedido.estado = "listo"
    pedido.save()

    channel_layer = get_channel_layer()

    if pedido.cliente:
        async_to_sync(channel_layer.group_send)(
            f"estado_pedidos_cliente_{pedido.cliente.id}",
            {
                "type": "actualizar_estado",
                "pedido": pedido.id,
                "estado": "listo"
            }
        )

    # 1) Actualizar en tiempo real (cocina + cajero)
    try:
        enviar_actualizacion_cocina(pedido)  # este manda al grupo pedidos_activos
    except Exception:
        # para que NO te salga “error de conexión” por WS si algo falla
        pass

    # 2) Notificación (mesero o cajero)
    destinatario = pedido.mesero if pedido.tipo_pedido == 'restaurante' else pedido.cajero
    if not destinatario:
        return JsonResponse({"success": False, "mensaje": "No se encontró el destinatario del pedido."})

    notif = Notificacion.objects.create(
        usuario_destino=destinatario,
        tipo="pedido_listo",
        mensaje=f"El pedido {pedido.codigo_pedido} está listo.",
        pedido=pedido
    )

    try:
        async_to_sync(channel_layer.group_send)(
            f"notificaciones_{destinatario.id}",
            {
                "type": "enviar_notificacion",
                "tipo": "pedido_listo",
                "mensaje": f"El pedido {pedido.codigo_pedido} está listo.",
                "mesa": pedido.mesa.numero if pedido.mesa else None,
                "pedido": pedido.id,
                "codigo_pedido": pedido.codigo_pedido,
                "id": notif.id,
                "fecha": localtime(notif.fecha_hora).strftime("%d/%m/%Y %H:%M"),
            }
        )

        if destinatario.rol == "mesero":
            url_destino = reverse("listar_pedidos")
        elif destinatario.rol == "cajero":
            url_destino = reverse("cajero_listar_pedidos")
        else:
            url_destino = "/"

        enviar_push(
            destinatario,
            "Pedido listo",
            f"El pedido {pedido.codigo_pedido} está listo.",
            url_destino
        )
    except Exception:
        pass

    # 3) Evento cobro (para que aparezca en "Pedidos a Cobrar")
    mesa_texto = pedido.mesa.numero if pedido.mesa else "Domicilio"

    if pedido.tipo_pedido == 'restaurante' and pedido.mesero:
        nombre_atendio = pedido.mesero.nombre
    elif pedido.tipo_pedido == 'domicilio' and pedido.cajero:
        nombre_atendio = f"Cajero: {pedido.cajero.nombre}"
    else:
        nombre_atendio = "N/A"

    # ESTE ES EL NOMBRE CORRECTO DEL CLIENTE
    nombre_cliente = pedido.cliente_nombre  # usa tu @property

    try:
        async_to_sync(channel_layer.group_send)(
            "pedidos_activos",
            {
                "type": "nuevo_cobro",
                "origen": "domicilio" if pedido.tipo_pedido == "domicilio" else "restaurante",
                "pedido_id": pedido.id,
                "codigo_pedido": pedido.codigo_pedido,
                "mesa": mesa_texto,
                "mesero": nombre_atendio,

                # cliente correcto
                "cliente": nombre_cliente,

                "total": float(pedido.total or 0),
                "estado_pago": "pendiente",

                # para filtrar en frontend y que no lo vean otros cajeros
                "cajero_id": pedido.cajero_id if pedido.tipo_pedido == "domicilio" else None,
            }
        )
    except Exception:
        pass


    return JsonResponse({"success": True})

def _payload_pedido_cocina(pedido):
    return {
        "id": pedido.id,
        "codigo_pedido": pedido.codigo_pedido,
        "tipo": pedido.tipo_pedido,
        "estado": pedido.estado,

        # restaurante
        "mesa": pedido.mesa.numero if pedido.mesa else None,
        "mesero": pedido.mesero.nombre if pedido.mesero else None,

        # domicilio (PARA LA TABLA DEL CAJERO)
        "nombre_cliente": pedido.nombre_cliente or "",
        "contacto_cliente": pedido.contacto_cliente or "",
        "total": float(pedido.total or 0),

        "cajero_id": pedido.cajero_id,

        "productos": [
            {"nombre": d.producto.nombre, "cantidad": d.cantidad, "observacion": d.observacion or ""}
            for d in pedido.detalles.select_related("producto").all()
        ]
    }

def enviar_pedido_cocina(pedido):
    if pedido.estado not in ["en_creacion", "aceptado", "en preparacion"]:
        return

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "nuevo_pedido",
            "pedido": _payload_pedido_cocina(pedido),
        }
    )


def enviar_actualizacion_cocina(pedido):

    # SOLO si ya fue enviado a cocina
    if not pedido.enviado_cocina:
        return

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "actualizar_pedido",
            "pedido": _payload_pedido_cocina(pedido),
        }
    )

@login_required(login_url='login')
@rol_requerido('cocinero')
@require_POST
@csrf_protect
def enviar_mensaje_mesero(request):
    data = json.loads(request.body)

    usuario_id = data.get("usuario_id")
    mensaje = (data.get("mensaje") or "").strip()

    if not usuario_id or not mensaje:
        return JsonResponse({"success": False, "mensaje": "Datos incompletos."}, status=400)

    destinatario = get_object_or_404(Usuario, id=usuario_id)

    # ahora request.user SIEMPRE será Usuario (porque login_required)
    msg = Mensaje.objects.create(
        remitente=request.user,
        destinatario=destinatario,
        contenido=mensaje
    )

    notif = Notificacion.objects.create(
        usuario_destino=destinatario,
        tipo="mensaje",
        mensaje=mensaje
    )

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f"notificaciones_{destinatario.id}",
        {
            "type": "enviar_notificacion",
            "tipo": "mensaje",
            "mensaje": mensaje,
            "id": notif.id,
            "fecha": notif.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            "pedido": None,
            "mesa": None,
        }
    )

    return JsonResponse({"success": True})

@login_required(login_url='login')
@rol_requerido('cocinero')
@require_POST
@csrf_protect
def marcar_pedido_preparacion(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado == "en preparacion":
        return JsonResponse({
            "success": False,
            "mensaje": "Este pedido ya está siendo preparado."
        })

    if pedido.estado not in ["en_creacion", "aceptado"]:
        return JsonResponse({
            "success": False,
            "mensaje": "Este pedido no puede pasar a preparación."
        })

    pedido.estado = "en preparacion"
    pedido.save(update_fields=["estado"])

    channel_layer = get_channel_layer()

    # SOLO estado (sin notificación)
    if pedido.cliente:
        async_to_sync(channel_layer.group_send)(
            f"estado_pedidos_cliente_{pedido.cliente.id}",
            {
                "type": "actualizar_estado",
                "pedido": pedido.id,
                "estado": "en preparacion"
            }
        )

    # WS cocina / cajero (ya lo tenías)
    try:
        enviar_actualizacion_cocina(pedido)
    except Exception:
        pass

    return JsonResponse({"success": True})

@login_required(login_url='login')
@rol_requerido('cocinero')
@require_POST
@csrf_protect
def avisar_no_hay_producto(request):
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "mensaje": "JSON inválido."}, status=400)

    producto_id = data.get("producto_id")
    if not producto_id:
        return JsonResponse({"success": False, "mensaje": "Falta producto_id."}, status=400)

    producto = Producto.objects.filter(id=producto_id).first()
    if not producto:
        return JsonResponse({"success": False, "mensaje": "Producto no encontrado."}, status=404)

    # Si ya está agotado hoy, no re-enviar (y mantiene “Enviado” al recargar)
    hoy = timezone.localdate()
    if producto.agotado_fecha == hoy:
        return JsonResponse({"success": True, "mensaje": f'Ya estaba marcado como agotado hoy: {producto.nombre}.'})

    # Marcar agotado HOY (persistente)
    producto.agotado_fecha = hoy
    producto.agotado_hora = timezone.now()
    producto.save(update_fields=["agotado_fecha", "agotado_hora"])

    destinatarios = Usuario.objects.filter(rol__in=["mesero", "cajero"]).only("id", "rol")
    if not destinatarios.exists():
        return JsonResponse({"success": False, "mensaje": "No hay destinatarios (meseros/cajeros)."}, status=404)

    mensaje = f'AGOTADO HOY: "{producto.nombre}". No ofrecer ni agregar a pedidos.'

    channel_layer = get_channel_layer()

    # Guardar notificaciones
    notifs = [Notificacion(usuario_destino=u, tipo="producto_no_hay", mensaje=mensaje) for u in destinatarios]
    Notificacion.objects.bulk_create(notifs)

    # Guardar mensajes (opcional)
    msgs = [Mensaje(remitente=request.user, destinatario=u, contenido=mensaje) for u in destinatarios]
    Mensaje.objects.bulk_create(msgs)

    # Enviar WS + incluir producto_id para bloquear en UI en tiempo real
    for u in destinatarios:
        try:
            async_to_sync(channel_layer.group_send)(
                f"notificaciones_{u.id}",
                {
                    "type": "enviar_notificacion",
                    "tipo": "producto_no_hay",
                    "mensaje": mensaje,
                    "producto_id": producto.id,
                    "producto_nombre": producto.nombre,
                    "pedido": None,
                    "mesa": None,
                }
            )

            if u.rol == "mesero":
                url_destino = reverse("dashboard_mesero")  # /pedidos/
            elif u.rol == "cajero":
                url_destino = reverse("dashboard_cajero")
            else:
                url_destino = "/"

            enviar_push(
                u,
                "Producto agotado",
                f'{producto.nombre} está agotado hoy.',
                url_destino
            )


        except Exception:
            pass

    return JsonResponse({"success": True, "mensaje": f'Aviso enviado: {producto.nombre}.'})

# ----------------------------
# Cajero - Pedidos a domicilio
# ----------------------------
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_listar_pedidos(request):
    hoy = timezone.localdate()

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user,
        cliente__isnull=True,
        fecha_hora__date=hoy
    ).exclude(estado='borrador').order_by('id')


    return render(request, 'cajero/pedidos/listar_pedidos.html', {'pedidos': pedidos})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_crear_pedido(request):
    """
    Crea el pedido a domicilio: SOLO datos del cliente.
    El recargo por producto se pedirá al agregar cada detalle.
    """
    if request.method == 'POST':
        nombre_cliente = request.POST.get('nombre_cliente', '').strip()
        contacto_cliente = request.POST.get('contacto_cliente', '').strip()
        direccion_entrega = request.POST.get('direccion_entrega', '').strip()

        # Validaciones básicas
        if not nombre_cliente or not contacto_cliente or not direccion_entrega:
            messages.error(request, "Todos los datos del cliente son obligatorios.")
            return render(request, 'cajero/pedidos/crear_pedido.html', {
                'user': request.user,
                'nombre_cliente': nombre_cliente,
                'contacto_cliente': contacto_cliente,
                'direccion_entrega': direccion_entrega,
            })

        # Crear pedido a domicilio (sin recargo, se usará recargo por detalle)
        pedido = Pedido.objects.create(
            cajero=request.user,
            tipo_pedido='domicilio',
            estado='borrador',
            nombre_cliente=nombre_cliente,
            contacto_cliente=contacto_cliente,
            direccion_entrega=direccion_entrega,
            # recargo_domicilio = 0 por defecto si existe el campo
        )

        return redirect('cajero_agregar_detalles', pedido_id=pedido.id)

    return render(request, 'cajero/pedidos/crear_pedido.html', {
        'user': request.user
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_agregar_detalles(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user,
        estado__in=['borrador', 'en_creacion']
    )
    productos = Producto.objects.filter(activo=True).order_by("nombre")
    
    # CALCULAR RECARGOS AQUi
    total_recargos = pedido.detalles.aggregate(
        total=Sum(F('recargo') * F('cantidad'))
    )['total'] or 0

    return render(request, 'cajero/pedidos/agregar_detalles.html', {
        'pedido': pedido,
        'productos': productos,
        'total_recargos': total_recargos
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_agregar_detalle_ajax(request, pedido_id):
    """
    Agrega productos al pedido a domicilio.
    VALIDACIÓN: si el producto ya existe en el pedido, NO se vuelve a agregar.
    """
    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))
        observacion = data.get("observacion", '')

        # VALIDAR: NO permitir agregar si está agotado hoy
        producto = get_object_or_404(Producto, id=producto_id)
        if producto.agotado_hoy:
            return JsonResponse({
                "success": False,
                "mensaje": f'El producto "{producto.nombre}" está AGOTADO hoy y no se puede agregar.'
            })


        recargo_unitario = float(data.get("recargo_unitario", 0))
        recargo_total = float(data.get("recargo_total", 0))  # (no es necesario guardarlo)

        # VALIDAR DUPLICADO
        if DetallePedido.objects.filter(pedido_id=pedido_id, producto_id=producto_id).exists():
            return JsonResponse({
                'success': False,
                'mensaje': 'Este producto ya fue agregado al pedido. Edítelo en la tabla.'
            })

        # CREAR DETALLE (SIN DUPLICAR)
        DetallePedido.objects.create(
            pedido_id=pedido_id,
            producto_id=producto_id,
            cantidad=cantidad,
            observacion=observacion,
            recargo=recargo_unitario
        )

        pedido = get_object_or_404(Pedido, id=pedido_id)
        pedido.calcular_totales()

        total_recargos = pedido.detalles.aggregate(
            total=Sum(F('recargo') * F('cantidad'))
        )['total'] or 0

        # ORDEN ORIGINAL (por id ascendente)
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('cajero/pedidos/tabla_detalles_acciones.html', {
            'pedido': pedido,
            'total_recargos': total_recargos,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Producto agregado con recargo.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Método no permitido.'})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_editar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if pedido.tipo_pedido != 'domicilio' or pedido.cajero != request.user:
        return JsonResponse({'success': False, 'mensaje': 'No autorizado.'}, status=403)

    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        detalle.cantidad = int(data.get("cantidad", detalle.cantidad))
        detalle.observacion = data.get("observacion", detalle.observacion)
        detalle.recargo = float(data.get("recargo", detalle.recargo))
        detalle.save()

        # recalcular totales del pedido
        pedido.calcular_totales()

        total_recargos = pedido.detalles.aggregate(
            total=Sum(F('recargo') * F('cantidad'))
        )['total'] or 0

        # ORDEN ORIGINAL
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('cajero/pedidos/tabla_detalles_acciones.html', {
            'pedido': pedido,
            'total_recargos': total_recargos,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Detalle actualizado correctamente.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Método no permitido.'})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_eliminar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if pedido.tipo_pedido != 'domicilio' or pedido.cajero != request.user:
        return JsonResponse({'success': False, 'mensaje': 'No autorizado.'}, status=403)

    if request.method == 'POST':
        detalle.delete()

        #  recalcular totales
        pedido.calcular_totales()

        total_recargos = pedido.detalles.aggregate(
            total=Sum(F('recargo') * F('cantidad'))
        )['total'] or 0

        #  ORDEN ORIGINAL
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string('cajero/pedidos/tabla_detalles_acciones.html', {
            'pedido': pedido,
            'total_recargos': total_recargos,
            'detalles': detalles
        })

        return JsonResponse({
            'success': True,
            'mensaje': 'Producto eliminado correctamente.',
            'tabla': tabla_html
        })

    return JsonResponse({'success': False, 'mensaje': 'Método no permitido.'})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_ver_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user
    )

    if not pedido.detalles.exists():
        messages.error(
            request,
            "No se puede enviar a cocina un pedido sin productos."
        )
        return redirect('cajero_agregar_detalles', pedido_id=pedido.id)

    # SOLO pedidos creados por cajero manualmente
    if (
        pedido.cliente is None
        and pedido.cajero == request.user
        and pedido.estado in ["borrador", "en_creacion"]
        and not pedido.enviado_cocina
    ):

        pedido.estado = "en_creacion"
        pedido.enviado_cocina = True
        pedido.save(update_fields=["estado", "enviado_cocina"])
        enviar_pedido_cocina(pedido)

   # solo una vez
    # si ya está en preparación/listo/finalizado, NO lo vuelves a mandar

    return render(request, 'cajero/pedidos/ver_pedido.html', {
        'pedido': pedido
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_api_estados_pedidos(request):
    hoy = timezone.localdate()

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user,
        cliente__isnull=True,
        fecha_hora__date=hoy
    ).exclude(estado='borrador').values(
        'id', 'estado', 'codigo_pedido', 'nombre_cliente', 'contacto_cliente', 'total'
    )

    return JsonResponse(list(pedidos), safe=False)

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user
    )

    if pedido.estado not in ['borrador', 'en_creacion']:
        return JsonResponse({
            "success": False,
            "message": "Solo se puede eliminar un pedido que aún no esté en preparación."
        })

    if request.method == 'POST':
        pedido_id = pedido.id
        pedido.delete()

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "pedidos_activos",
            {
                "type": "eliminar_pedido",
                "pedido_id": pedido_id
            }
        )

        return JsonResponse({"success": True})

    return JsonResponse({"success": False})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_finalizar_pedido(request, pedido_id):

    if request.method != 'POST':
        return JsonResponse({"success": False, "message": "Método no permitido."})

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user
    )

    if pedido.estado != "listo":
        return JsonResponse({
            "success": False,
            "message": "Solo se puede finalizar un pedido que ya esté listo."
        })

    # CAMBIO DE ESTADO
    pedido.estado = "finalizado"
    pedido.save(update_fields=["estado"])

    # AVISAR EN TIEMPO REAL AL CAJERO
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "actualizar_pedido",
            "pedido": payload_pedido_cajero(pedido)
        }
    )

    return JsonResponse({
        "success": True,
        "message": f"El pedido #{pedido.codigo_pedido} ha sido finalizado."
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_editar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user
    )

    # NECESARIOS PARA QUE SE VEAN LOS PRODUCTOS Y DETALLES
    productos = Producto.objects.filter(activo=True).order_by("nombre")
    detalles = pedido.detalles.all()

    if request.method == "POST":

        nombre_cliente = request.POST.get('nombre_cliente', '').strip()
        contacto_cliente = request.POST.get('contacto_cliente', '').strip()
        direccion_entrega = request.POST.get('direccion_entrega', '').strip()
        recargo_domicilio = request.POST.get('recargo_domicilio', '0').strip()

        # Validaciones
        if not nombre_cliente or not contacto_cliente or not direccion_entrega:
            return JsonResponse({
                "success": False,
                "mensaje": "Todos los campos son obligatorios."
            })

        try:
            recargo_decimal = float(recargo_domicilio)
        except:
            return JsonResponse({
                "success": False,
                "mensaje": "El recargo debe ser un número válido."
            })

        # Guardar cambios
        pedido.nombre_cliente = nombre_cliente
        pedido.contacto_cliente = contacto_cliente
        pedido.direccion_entrega = direccion_entrega
        pedido.recargo_domicilio = recargo_decimal
        pedido.save()
        pedido.calcular_totales()

        # SOLO si el pedido fue creado por cajero
        if pedido.cliente is None:
            enviar_actualizacion_cocina(pedido)



        return JsonResponse({
            "success": True,
            "mensaje": f"Pedido #{pedido.id} actualizado correctamente.",
            "pedido_id": pedido.id
        })

    # SE ENVÍAN LOS PRODUCTOS Y DETALLES AQUÍ
    return render(request, 'cajero/pedidos/editar_pedido.html', {
        'pedido': pedido,
        'productos': productos,    
        'detalles': detalles      
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_cancelar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio',
        cajero=request.user,
        estado='borrador'
    )

    pedido.delete()

    messages.info(request, "Pedido cancelado correctamente.")
    return redirect('cajero_listar_pedidos')

def pedido_to_dict(pedido):
    return {
        "id": pedido.id,
        "codigo_pedido": pedido.codigo_pedido,
        "tipo": pedido.tipo_pedido,
        "estado": pedido.estado,

        "mesa": pedido.mesa.numero if pedido.mesa else None,
        "mesero": pedido.mesero.nombre if pedido.mesero else None,

        "nombre_cliente": pedido.nombre_cliente or "",
        "contacto_cliente": pedido.contacto_cliente or "",
        "total": float(pedido.total or 0),

        "cajero": pedido.cajero.nombre if pedido.cajero else None,

        "productos": [
            {"nombre": d.producto.nombre, "cantidad": d.cantidad, "observacion": d.observacion or ""}
            for d in pedido.detalles.select_related("producto").all()
        ]
    }

def payload_pedido_cajero(pedido):
    return {
        "id": pedido.id,
        "codigo_pedido": pedido.codigo_pedido,
        "tipo": pedido.tipo_pedido,
        "estado": pedido.estado,   # en_creacion | en preparacion | listo | finalizado
        "nombre_cliente": pedido.nombre_cliente or "",
        "contacto_cliente": pedido.contacto_cliente or "",
        "total": float(pedido.total or 0),
        "cajero_id": pedido.cajero_id,
    }

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_historial_pedidos(request):

    codigo = request.GET.get("codigo", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    metodo = request.GET.get("metodo", "")

    pedidos = (
        Pedido.objects
        .filter(
            tipo_pedido='domicilio',
            cajero=request.user,
            cliente__isnull=True,
            estado='finalizado',
            pagos__estado_pago='confirmado'
        )
        .prefetch_related(
            Prefetch(
                'pagos',
                queryset=Pago.objects.filter(estado_pago='confirmado')
                .prefetch_related('comprobante')
            )
        )
        .distinct()
        .order_by('-fecha_hora')
    )

    # =======================
    # FILTROS
    # =======================
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo:
        pedidos = pedidos.filter(pagos__metodo_pago=metodo)

    # 🔹 TOTAL RECAUDADO
    total_recaudado = pedidos.aggregate(total=Sum('total'))['total'] or 0

    # Pago y comprobante
    for p in pedidos:
        p.pago_confirmado = p.pagos.first()

        if p.pago_confirmado:
            p.comprobante = p.pago_confirmado.comprobante.first()
        else:
            p.comprobante = None

    return render(
        request,
        'cajero/pedidos/historial_pedidos.html',
        {
            'pedidos': pedidos,
            'total_recaudado': total_recaudado,
            'filtros': {
                'codigo': codigo,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'metodo': metodo,
            }
        }
    )

@login_required(login_url='login')
@rol_requerido('cajero')
def exportar_historial_cobrados_excel(request):

    codigo = request.GET.get("codigo", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    metodo = request.GET.get("metodo", "")

    pedidos = (
        Pedido.objects
        .filter(
            tipo_pedido='domicilio',
            cajero=request.user,
            cliente__isnull=True,
            estado='finalizado',
            pagos__estado_pago='confirmado'
        )
        .prefetch_related(
            Prefetch(
                'pagos',
                queryset=Pago.objects.filter(estado_pago='confirmado')
            )
        )
        .distinct()
        .order_by('-fecha_hora')
    )

    # FILTROS
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo:
        pedidos = pedidos.filter(pagos__metodo_pago=metodo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Cobrados"

    headers = ["Código", "Cliente", "Total", "Método", "Fecha"]
    ws.append(headers)

    for p in pedidos:
        pago = p.pagos.first()
        metodo_pago = pago.metodo_pago.capitalize() if pago else ""

        ws.append([
            p.codigo_pedido,
            p.nombre_cliente,
            float(p.total),
            metodo_pago,
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pedidos_cobrados.xlsx"'

    wb.save(response)

    return response

@login_required(login_url='login')
@rol_requerido('cajero')
def exportar_historial_cobrados_pdf(request):

    codigo = request.GET.get("codigo", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    metodo = request.GET.get("metodo", "")

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user,
        cliente__isnull=True,
        estado='finalizado',
        pagos__estado_pago='confirmado'
    ).prefetch_related('pagos').distinct().order_by('-fecha_hora')

    # FILTROS
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo:
        pedidos = pedidos.filter(pagos__metodo_pago=metodo)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos_cobrados.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    elementos.append(Paragraph(
        "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO",
        ParagraphStyle(name='Titulo', alignment=1, fontSize=14)
    ))

    elementos.append(Spacer(1,10))

    elementos.append(Paragraph(
        f"Reporte generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle(name='Fecha', alignment=1)
    ))

    elementos.append(Spacer(1,15))

    data = [["Código", "Cliente", "Total", "Método", "Fecha"]]

    for p in pedidos:
        pago = p.pagos.first()
        metodo_pago = pago.metodo_pago.capitalize() if pago else ""

        data.append([
            p.codigo_pedido,
            p.nombre_cliente,
            f"$ {p.total:.2f}",
            metodo_pago,
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[80,120,70,90,110])

    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    return response


#-------------------------------
#  CAJERO - PEDIDOS CLIENTE
# ------------------------------
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_pedidos_clientes_domicilio(request):

    hoy = timezone.localdate()

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        fecha_hora__date=hoy,
        detalles__isnull=False,
        comprobantes_cliente__estado='pendiente'
    ).filter(
        Q(estado='pendiente_caja', cajero__isnull=True) 
    ).distinct().order_by('id')
    

    return render(
        request,
        'cajero/pedidos/pedidos_clientes_domicilio.html',
        {'pedidos': pedidos}
    )


@login_required(login_url='login')
@rol_requerido('cajero')
@require_POST
def cajero_aceptar_pedido_cliente(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio'
    )

    if pedido.estado != 'pendiente_caja':
        return JsonResponse({
            "success": False,
            "message": "Este pedido ya fue procesado."
        })

    # ================================
    # 1) ACEPTAR PEDIDO
    # ================================
    pedido.cajero = request.user
    pedido.estado = 'aceptado'
    pedido.enviado_cocina = True
    pedido.save(update_fields=['cajero', 'estado', 'enviado_cocina'])

    # ================================
    # DEFINIR DATOS DEL COMPROBANTE
    # ================================
    numero = f"DC-{pedido.id}"  # o el formato que tú quieras

    # Para pedidos de cliente (domicilio)
    nombre_comp = pedido.cliente_nombre
    direccion_comp = pedido.direccion_entrega or ""

    # ================================
    # 2) CREAR PAGO + COMPROBANTE
    # ================================
    with transaction.atomic():
        pago = Pago.objects.create(
            pedido=pedido,
            total=Decimal(str(pedido.total)),
            monto_recibido=Decimal(str(pedido.total)),
            metodo_pago="transferencia",
            estado_pago="confirmado"
        )

        try:
            comprobante = Comprobante.objects.create(
                pago=pago,
                numero_comprobante=numero,
                nombre_cliente=nombre_comp,
                direccion_cliente=direccion_comp,
                correo_cliente=None
            )
        except IntegrityError:
            return JsonResponse({
                "success": False,
                "message": "El número de comprobante ya existe. Verifique e intente nuevamente."
            })


    generar_comprobante_pdf(request, comprobante)
    comprobante.refresh_from_db()

    import cloudinary.utils

    comprobante_url = ""

    if comprobante.archivo_pdf:
        comprobante_url = comprobante.archivo_pdf.url

    # ================================
    #  DEFINIR UNA SOLA VEZ
    # ================================
    channel_layer = get_channel_layer()

    # ================================
    # 3) WS → ACTUALIZAR ESTADO CLIENTE
    # ================================
    async_to_sync(channel_layer.group_send)(
        f"estado_pedidos_cliente_{pedido.cliente.id}",
        {
            "type": "actualizar_estado",
            "pedido": pedido.id,
            "estado": "aceptado",
            "comprobante_url": comprobante_url
        }
    )

    # ================================
    # 4) NOTIFICACIÓN CLIENTE
    # ================================
    notif = Notificacion.objects.create(
        pedido=pedido,
        usuario_destino=pedido.cliente,
        tipo="comprobante_generado",
        mensaje=f"Tu pedido #{pedido.codigo_pedido} fue aceptado."
    )

    async_to_sync(channel_layer.group_send)(
        f"notificaciones_{pedido.cliente.id}",
        {
            "type": "enviar_notificacion",
            "mensaje": "Tu pedido fue aceptado.",
            "pedido": pedido.id,
            "comprobante_url": comprobante_url,
            "fecha": localtime(notif.fecha_hora).strftime("%d/%m/%Y %H:%M"),
        }
    )

    enviar_push(
        pedido.cliente,
        "Pedido aceptado",
        f"Tu pedido #{pedido.codigo_pedido} fue aceptado.",
        reverse("cliente_listar_pedidos")
    )

    # ================================
    # 5) ENVIAR A COCINA
    # ================================
    enviar_pedido_cocina(pedido)

    return JsonResponse({'success': True})

@login_required(login_url='login')
@rol_requerido('cajero')
@require_POST
def cajero_rechazar_pedido_cliente(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        tipo_pedido='domicilio'
    )

    if pedido.estado != 'pendiente_caja':
        return JsonResponse({
            "success": False,
            "message": "Este pedido ya fue procesado."
        })

    comprobante = pedido.comprobantes_cliente.filter(estado='pendiente').first()
    if comprobante:
        comprobante.estado = 'rechazado'
        comprobante.save(update_fields=['estado'])

    pedido.estado = 'rechazado'
    pedido.enviado_cocina = False
    pedido.cajero = request.user
    pedido.save(update_fields=['estado', 'enviado_cocina', 'cajero'])

    channel_layer = get_channel_layer()

    # WS actualizar estado cliente
    async_to_sync(channel_layer.group_send)(
        f"estado_pedidos_cliente_{pedido.cliente.id}",
        {
            "type": "actualizar_estado",
            "pedido": pedido.id,
            "estado": "rechazado"
        }
    )

    # Notificación DB
    notif = Notificacion.objects.create(
        pedido=pedido,
        usuario_destino=pedido.cliente,
        tipo="pedido_rechazado",
        mensaje=f"Tu pedido #{pedido.codigo_pedido} fue rechazado."
    )

    # PUSH REAL
    enviar_push(
        pedido.cliente,
        "Pedido rechazado",
        f"Tu pedido #{pedido.codigo_pedido} fue rechazado.",
        reverse("cliente_listar_pedidos")
    )

    return JsonResponse({'success': True})

@login_required
@rol_requerido('cajero')
def cajero_enviar_cocina(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        estado='aceptado',
        enviado_cocina=False
    )

    pedido.enviado_cocina = True
    pedido.estado = 'en preparacion'
    pedido.save(update_fields=['enviado_cocina','estado'])

    # notificación cocina aquí

    return JsonResponse({'success': True})

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_pedidos_historial_cliente(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user
    )

    estado = request.GET.get('estado')
    if estado in ['listo', 'rechazado']:
        pedidos = pedidos.filter(estado=estado)
    else:
        pedidos = pedidos.filter(estado__in=['listo', 'rechazado'])

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    codigo = request.GET.get('codigo')
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    pedidos = pedidos.order_by('-fecha_hora')

    # TOTAL RECAUDADO
    total_recaudado = pedidos.aggregate(total=Sum('total'))['total'] or 0

    return render(
        request,
        'cajero/pedidos/pedidos_aceptados_rechazados.html',
        {
            'pedidos': pedidos,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'codigo': codigo,
            'estado': estado,
            'total_recaudado': total_recaudado
        }
    )

@login_required(login_url='login')
@rol_requerido('cajero')
def exportar_historial_cliente_excel(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user
    ).select_related('cliente')

    # ===== FILTROS =====
    estado = request.GET.get('estado')
    if estado in ['listo', 'rechazado']:
        pedidos = pedidos.filter(estado=estado)
    else:
        pedidos = pedidos.filter(estado__in=['listo', 'rechazado'])

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    codigo = request.GET.get('codigo')
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    pedidos = pedidos.order_by('-fecha_hora')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Cliente"

    headers = ["Código", "Cliente", "Dirección", "Total", "Estado", "Fecha"]
    ws.append(headers)

    for p in pedidos:
        ws.append([
            p.codigo_pedido,
            p.cliente.nombre if p.cliente else "",
            p.direccion_entrega,
            float(p.total),
            p.estado.capitalize(),
            p.fecha_hora.strftime('%d/%m/%Y %H:%M')
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_clientes.xlsx"'

    wb.save(response)

    return response

@login_required(login_url='login')
@rol_requerido('cajero')
def exportar_historial_cliente_pdf(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cajero=request.user
    ).select_related('cliente')

    # ===== FILTROS =====
    estado = request.GET.get('estado')
    if estado in ['listo', 'rechazado']:
        pedidos = pedidos.filter(estado=estado)
    else:
        pedidos = pedidos.filter(estado__in=['listo', 'rechazado'])

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    codigo = request.GET.get('codigo')
    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    pedidos = pedidos.order_by('-fecha_hora')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historial_clientes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    elementos.append(Paragraph(
        "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO",
        ParagraphStyle(name='Titulo', alignment=1, fontSize=14, leading=16)
    ))

    elementos.append(Spacer(1,10))

    elementos.append(Paragraph(
        f"Reporte generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle(name='Fecha', alignment=1)
    ))

    elementos.append(Spacer(1,15))

    data = [["Código", "Cliente", "Dirección", "Total", "Estado", "Fecha"]]

    for p in pedidos:
        data.append([
            p.codigo_pedido,
            p.cliente.nombre if p.cliente else "",
            p.direccion_entrega,
            f"$ {p.total:.2f}",
            p.estado.capitalize(),
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[80,100,120,60,70,90])

    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    return response
#-------------------------------
# CAJERO - COBROS
#-------------------------------
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_restaurante_cobros(request):

    pagos_confirmados = Pago.objects.filter(
        pedido=OuterRef('pk'),
        estado_pago='confirmado'
    )

    hoy = timezone.localdate()

    # ---------- PEDIDOS PENDIENTES DE COBRO (SOLO HOY) ----------
    pedidos_cobrar = (
        Pedido.objects
        .filter(
            tipo_pedido='restaurante',
            estado__in=['listo', 'finalizado'],
            fecha_hora__date=hoy
        )
        .annotate(tiene_pago=Exists(pagos_confirmados))
        .filter(tiene_pago=False)
        .order_by('id')
    )

    # ---------- PEDIDOS PAGADOS (SOLO HOY) + COMPROBANTES ----------
    pedidos_pagados = (
        Pedido.objects
        .filter(
            tipo_pedido='restaurante',
            pagos__estado_pago='confirmado',
            pagos__fecha_hora__date=hoy
        )
        .prefetch_related('pagos__comprobante') # para acceder a los PDFs
        .distinct()
        .order_by('id')
    )

    # añadir el último pago confirmado a cada pedido (igual que domicilio)
    for p in pedidos_pagados:
        p.pago_confirmado = (
            p.pagos
             .filter(estado_pago='confirmado')
             .order_by('id')
             .first()
        )

    return render(request, "cajero/pago/cobros_restaurante.html", {
        "pedidos_cobrar": pedidos_cobrar,
        "pedidos_pagados": pedidos_pagados
    })

# ============================================================
#                    REGISTRAR PAGO DEL PEDIDO
# ============================================================

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_restaurante_pagar(request, pedido_id):

    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método no permitido."})

    pedido = get_object_or_404(Pedido, id=pedido_id, tipo_pedido="restaurante")

    # Solo permitir cobrar pedidos listos o finalizados
    if pedido.estado not in ["listo", "finalizado"]:
        return JsonResponse({"success": False, "message": "El pedido no puede ser pagado."})

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "message": "Datos inválidos."})

    metodo = (data.get("metodo") or "").strip()
    recibido_raw = data.get("recibido", None)
    referencia = (data.get("referencia") or "").strip()

    cliente_nombre = (data.get("cliente_nombre") or "").strip()
    cliente_direccion = (data.get("cliente_direccion") or "").strip()

    # Convertir a Decimal seguro
    try:
        total = Decimal(str(pedido.total)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if recibido_raw in ["", None]:
            recibido = None
        else:
            recibido = Decimal(str(recibido_raw)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return JsonResponse({"success": False, "message": "El monto ingresado no es válido."})

    # regla >= 50
    if total >= Decimal("50.00"):
        if not cliente_nombre or not cliente_direccion:
            return JsonResponse({
                "success": False,
                "message": "Total >= $50: debe ingresar Nombre y Dirección del cliente."
            })
    # Validaciones por método
    if metodo == "transferencia":
        if recibido is None:
            return JsonResponse({"success": False, "message": "Debe ingresar el monto de la transferencia."})
        if recibido != total:
            return JsonResponse({"success": False, "message": "El monto debe ser EXACTO para transferencias."})
        if not referencia:
            return JsonResponse({"success": False, "message": "Debe ingresar el número de comprobante."})
        cambio = Decimal("0.00")

    elif metodo == "efectivo":
        if recibido is None:
            return JsonResponse({"success": False, "message": "Debe ingresar el monto recibido."})
        if recibido < total:
            return JsonResponse({"success": False, "message": "El cliente no puede pagar menos del total."})
        cambio = (recibido - total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    else:
        return JsonResponse({"success": False, "message": "Método de pago no válido."})

    # ================================
    # VALIDAR NÚMERO DE COMPROBANTE ÚNICO
    # ================================
    if metodo == "transferencia":
        if Comprobante.objects.filter(numero_comprobante=referencia).exists():
            return JsonResponse({
                "success": False,
                "message": "El número de comprobante ya fue utilizado."
            })

    # ==========================================================
    # GUARDADO EN BD (atómico) + liberar mesa
    # ==========================================================
    with transaction.atomic():

        # Registrar pago
        pago = Pago.objects.create(
            pedido=pedido,
            total=total,
            monto_recibido=recibido,
            metodo_pago=metodo,
            referencia_transferencia=referencia if metodo == "transferencia" else "",
            cambio=cambio,
            estado_pago="confirmado"
        )

        # Finalizar pedido (si no lo está)
        if pedido.estado != "finalizado":
            pedido.estado = "finalizado"
            pedido.save(update_fields=["estado"])

        #  LIBERAR MESA (solo restaurante)
        if pedido.mesa_id:
            Mesa.objects.filter(id=pedido.mesa_id).update(estado="libre")

        # Comprobante
        if metodo == "transferencia":
            numero = referencia
        else:
            numero = f"C-{pedido.id}-{pago.id}"


        if total >= Decimal("50.00"):
            nombre_comp = cliente_nombre
            direccion_comp = cliente_direccion
        else:
            nombre_comp = "Consumidor final"
            direccion_comp = ""

        try:
            comprobante = Comprobante.objects.create(
                pago=pago,
                numero_comprobante=numero,
                nombre_cliente=nombre_comp,
                direccion_cliente=direccion_comp,
                correo_cliente=None
            )
        except IntegrityError:
            return JsonResponse({
                "success": False,
                "message": "El número de comprobante ya fue utilizado. Intente nuevamente."
            })

    # Generar PDF (fuera del atomic por si tarda)
    comprobante_url = ""
    try:
        generar_comprobante_pdf(request, comprobante)
        comprobante.refresh_from_db()
        if comprobante.archivo_pdf:
            comprobante_url = comprobante.archivo_pdf.url
        else:
            comprobante_url = ""

    except Exception as e:
        # IMPORTANTE: no rompas el cobro por el PDF
        print("ERROR generando PDF comprobante:", e)
        comprobante_url = ""  # quedará "No disponible" pero el cobro y WS sí salen
        
    # ==========================================================
    # WEBSOCKETS
    # ==========================================================
    channel_layer = get_channel_layer()

    # 1) SOLO CAJERO: quitarlo de "Pedidos a Cobrar"
    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "eliminar_pedido",
            "origen": "restaurante",
            "pedido_id": pedido.id,
            "accion": "quitar_de_cobros"
        }
    )

    # 2) MESERO (y cualquiera que muestre pedidos): actualizar estado a FINALIZADO
    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "actualizar_pedido",
            "pedido": {
                "id": pedido.id,
                "estado": "finalizado",
                "tipo": "restaurante",
            }
        }
    )

    # 3) Cajero: tabla "Pagados"
    fecha_local = timezone.localtime(pago.fecha_hora).strftime("%d/%m/%Y %H:%M")

    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "nuevo_pagado",
            "origen": "restaurante",
            "pedido_id": pedido.id,
            "codigo_pedido": pedido.codigo_pedido,
            "mesa": pedido.mesa.numero if pedido.mesa else "N/A",
            "mesero": pedido.mesero.nombre if pedido.mesero else "",
            "total": float(total),
            "fecha": fecha_local,
            "estado_pago": "confirmado",
            "comprobante_url": comprobante_url,
        }
    )

    return JsonResponse({
        "success": True,
        "message": "Pago registrado correctamente.",
        "comprobante_url": comprobante_url,
        "codigo_pedido": pedido.codigo_pedido
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_domicilio_cobros(request):

    hoy = timezone.localdate()

    pagos_confirmados = Pago.objects.filter(
        pedido=OuterRef('pk'),
        estado_pago='confirmado'
    )

    pedidos_cobrar = (
        Pedido.objects
        .filter(
            tipo_pedido='domicilio',
            cajero=request.user,
            fecha_hora__date=hoy,

            # SOLO pedidos creados por cajero (NO cliente)
            cliente__isnull=True,

            estado__in=['listo', 'finalizado']
        )
        .annotate(tiene_pago=Exists(pagos_confirmados))
        .filter(tiene_pago=False)
        .order_by('id')
    )

    pedidos_pagados = (
        Pedido.objects
        .filter(
            tipo_pedido='domicilio',
            cajero=request.user,
            pagos__estado_pago='confirmado',
            pagos__fecha_hora__date=hoy
        )
        .prefetch_related('pagos__comprobante')
        .distinct()
        .order_by('id')
    )

    for p in pedidos_pagados:
        p.pago_confirmado = (
            p.pagos.filter(estado_pago='confirmado').order_by('id').first()
        )

    return render(request, "cajero/pago/cobros_domicilio.html", {
        "pedidos_cobrar": pedidos_cobrar,
        "pedidos_pagados": pedidos_pagados
    })

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_domicilio_pagar(request, pedido_id):

    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método no permitido."})

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "message": "Datos inválidos."})

    metodo = data.get("metodo")
    recibido = data.get("recibido")
    referencia = (data.get("referencia") or "").strip()

    try:
        pedido = Pedido.objects.get(
            id=pedido_id,
            tipo_pedido="domicilio",
            cajero=request.user
        )
    except Pedido.DoesNotExist:
        return JsonResponse({"success": False, "message": "Pedido no encontrado."})

    total = float(pedido.total or 0)

    # ================================
    # VALIDACIÓN DE MÉTODO DE PAGO
    # ================================
    if metodo not in ["efectivo", "transferencia"]:
        return JsonResponse({"success": False, "message": "Método de pago inválido."})

    if metodo == "efectivo":
        try:
            recibido = float(str(recibido).replace(",", "."))
        except Exception:
            return JsonResponse({"success": False, "message": "Monto recibido inválido."})

        if recibido < total:
            return JsonResponse({"success": False, "message": "El monto recibido es insuficiente."})

    if metodo == "transferencia":
        if not referencia:
            return JsonResponse({"success": False, "message": "Debe ingresar número de comprobante."})

        try:
            recibido = float(str(recibido).replace(",", "."))
        except Exception:
            return JsonResponse({"success": False, "message": "Monto transferido inválido."})

        # Si quieres EXACTO igual al total, cambia por: if recibido != total:
        if recibido < total:
            return JsonResponse({"success": False, "message": "El monto transferido es insuficiente."})

    # Fecha local
    fecha_local = timezone.localtime(timezone.now())

    if metodo == "transferencia":
        if Comprobante.objects.filter(numero_comprobante=referencia).exists():
            return JsonResponse({
                "success": False,
                "message": "El número de comprobante ya fue utilizado."
            })

    # ================================
    # REGISTRO DEL PAGO
    # ================================
    with transaction.atomic():

        pago = Pago.objects.create(
            pedido=pedido,
            total=total,
            metodo_pago=metodo,
            monto_recibido=recibido,
            cambio=recibido - total if metodo == "efectivo" else 0,
            referencia_transferencia=referencia if metodo == "transferencia" else "",
            estado_pago="confirmado",
            fecha_hora=fecha_local,
        )

        # Cambiar estado del pedido (si así lo manejas en domicilio)
        pedido.estado = "finalizado"
        pedido.save(update_fields=["estado"])

    # ================================
    # COMPROBANTE
    # ================================
    numero = f"D-{pedido.id}-{pago.id}"

    try:
        comprobante = Comprobante.objects.create(
            pago=pago,
            numero_comprobante=numero,
            nombre_cliente="Consumidor final",
            direccion_cliente=pedido.direccion_entrega,
            correo_cliente=None
        )
    except IntegrityError:
        return JsonResponse({
            "success": False,
            "message": "El número de comprobante ya fue utilizado. Intente nuevamente."
        })


    # ================================
    # GENERAR PDF
    # ================================
    comprobante_url = ""

    try:
        generar_comprobante_pdf(request, comprobante)
        comprobante.refresh_from_db()

        if comprobante.archivo_pdf:
            comprobante_url = comprobante.archivo_pdf.url
        else:
            comprobante_url = ""

    except Exception as e:
        import traceback
        print("ERROR generando PDF comprobante:")
        traceback.print_exc()
        comprobante_url = ""

    # ================================
    # WEBSOCKET → NOTIFICAR TABLA PAGADOS
    # ================================
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        "pedidos_activos",
        {
            "type": "nuevo_pagado",
            "origen": "domicilio",
            "pedido_id": pedido.id,
            "cliente": pedido.nombre_cliente,
            "total": float(total),
            "fecha": fecha_local.strftime("%d/%m/%Y %H:%M"),
            "estado_pago": "confirmado",
            "comprobante_url": comprobante_url,
        }
    )

    # ================================
    # RESPUESTA
    # ================================
    return JsonResponse({
        "success": True,
        "message": "Pago registrado correctamente.",
        "comprobante_url": comprobante_url
    })

@login_required(login_url='login')
def ver_comprobante(request, comp_id):
    comprobante = get_object_or_404(Comprobante, id=comp_id)
    pago = comprobante.pago
    pedido = pago.pedido
    detalles = pedido.detalles.all()

    requiere_datos = pedido.total >= Decimal("50.00")

    # Datos sugeridos
    nombre_impreso = pedido.cliente_nombre  # ya te devuelve nombre de FK o nombre_cliente
    direccion_impresa = (
        (pedido.direccion_entrega or "").strip() or
        (pedido.cliente.direccion.strip() if pedido.cliente and pedido.cliente.direccion else "")
    )

    if pedido.tipo_pedido == "restaurante":
        template = "cajero/comprobantes/comprobante_restaurante.html"
    else:
        template = "cajero/comprobantes/comprobante_domicilio.html"

    return render(request, template, {
        "comprobante": comprobante,
        "pago": pago,
        "pedido": pedido,
        "detalles": detalles,
        "requiere_datos": requiere_datos,
        "nombre_impreso": nombre_impreso,
        "direccion_impresa": direccion_impresa,
    })

def generar_comprobante_pdf(request, comprobante):
    pago = comprobante.pago
    pedido = pago.pedido
    detalles = pedido.detalles.all()

    # ================================
    # VALIDACIÓN DE DATOS OBLIGATORIOS
    # ================================
    requiere_datos = pago.total >= Decimal("50.00")

    nombre_impreso = (comprobante.nombre_cliente or "").strip()
    direccion_impresa = (comprobante.direccion_cliente or "").strip()

    if requiere_datos:
        if not nombre_impreso or not direccion_impresa:
            raise ValueError(
                "Pedido >= $50 requiere nombre y dirección del cliente."
            )
    else:
        nombre_impreso = "Consumidor final"
        direccion_impresa = ""

    comprobante.nombre_cliente = nombre_impreso
    comprobante.direccion_cliente = direccion_impresa
    comprobante.save(update_fields=["nombre_cliente", "direccion_cliente"])

    # ================================
    # SELECCIÓN DE TEMPLATE
    # ================================
    if pedido.tipo_pedido == "restaurante":
        template = "cajero/comprobantes/comprobante_restaurante.html"

    elif pedido.tipo_pedido == "domicilio" and pedido.cliente is not None:
        template = "cajero/comprobantes/comprobante_domicilio_cliente.html"

    else:
        template = "cajero/comprobantes/comprobante_domicilio.html"

    # ================================
    # RENDER HTML
    # ================================
    html_string = render_to_string(template, {
        "comprobante": comprobante,
        "pago": pago,
        "pedido": pedido,
        "detalles": detalles,
        "requiere_datos": requiere_datos,
        "nombre_impreso": nombre_impreso,
        "direccion_impresa": direccion_impresa,
    })

    # ================================
    # GENERAR PDF
    # ================================
    pdf_bytes = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    nombre_archivo = f"{comprobante.numero_comprobante}.pdf"

    # ================================
    # SUBIR A CLOUDINARY
    # ================================
    import io
    import cloudinary
    import cloudinary.uploader

    pdf_file = io.BytesIO(pdf_bytes)
    pdf_file.name = f"{comprobante.numero_comprobante}.pdf"

    resultado = cloudinary.uploader.upload(
        pdf_file,
        resource_type="raw",
        folder="comprobantes",
        public_id=comprobante.numero_comprobante,
        overwrite=True
    )

    # Guardar SOLO el public_id
    comprobante.archivo_pdf = resultado["public_id"]
    comprobante.save(update_fields=["archivo_pdf"])
#-----------------------------
# Reportes
#-----------------------------
def reportes_general(request):
    return render(request, 'administrador/reportes/reportes_general.html')

@login_required(login_url='login')
@rol_requerido('admin')
def reporte_pagos_restaurante(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='restaurante',
        estado_pago='confirmado'
    ).select_related('pedido')

    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")
    metodo = request.GET.get("metodo")
    mesero_id = request.GET.get("mesero")

    if fecha_inicio:
        pagos = pagos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pagos = pagos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo and metodo != "todos":
        pagos = pagos.filter(metodo_pago=metodo)

    if mesero_id and mesero_id != "todos":
        pagos = pagos.filter(pedido__mesero_id=mesero_id)

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    meseros = Usuario.objects.filter(rol='mesero')

    return render(request, "administrador/reportes/reporte_restaurante.html", {
        "pagos": pagos,
        "total_recaudado": total_recaudado,
        "meseros": meseros,
    })

@login_required(login_url='login')
@rol_requerido('admin')
def reporte_pagos_domicilio(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='domicilio',
        estado_pago='confirmado'
    ).select_related('pedido')

    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")
    metodo = request.GET.get("metodo")
    cajero_id = request.GET.get("cajero")

    if fecha_inicio:
        pagos = pagos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pagos = pagos.filter(fecha_hora__date__lte=fecha_fin)

    if metodo and metodo != "todos":
        pagos = pagos.filter(metodo_pago=metodo)

    if cajero_id and cajero_id != "todos":
        pagos = pagos.filter(pedido__cajero_id=cajero_id)

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    cajeros = Usuario.objects.filter(rol='cajero')

    return render(request, "administrador/reportes/reporte_domicilio.html", {
        "pagos": pagos,
        "total_recaudado": total_recaudado,
        "cajeros": cajeros,
    })

@login_required(login_url='login')
@rol_requerido('admin')
def reporte_pedidos_restaurante(request):

    pedidos = (
        Pedido.objects
        .filter(tipo_pedido='restaurante')
        .select_related('mesa', 'mesero')
        .prefetch_related('detalles')
    )

    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")
    estado = request.GET.get("estado")
    mesero_id = request.GET.get("mesero")

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if estado and estado != "todos":
        pedidos = pedidos.filter(estado=estado)

    if mesero_id and mesero_id != "todos":
        pedidos = pedidos.filter(mesero_id=mesero_id)

    total_pedidos = pedidos.count()
    total_ventas = pedidos.aggregate(total=Sum('total'))['total'] or 0

    meseros = Usuario.objects.filter(rol='mesero')

    return render(request, "administrador/reportes/reporte_pedidos_restaurante.html", {
        "pedidos": pedidos.order_by('-fecha_hora'),
        "total_pedidos": total_pedidos,
        "total_ventas": total_ventas,
        "meseros": meseros,
    })

@login_required(login_url='login')
@rol_requerido('admin')
def reporte_pedidos_domicilio(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio'
    ).select_related('cajero')

    # Filtros
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")
    estado = request.GET.get("estado")
    cajero_id = request.GET.get("cajero")

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fecha_fin)

    if estado and estado != "todos":
        pedidos = pedidos.filter(estado=estado)

    if cajero_id and cajero_id != "todos":
        pedidos = pedidos.filter(cajero_id=cajero_id)

    total_pedidos = pedidos.count()
    total_ventas = pedidos.aggregate(total=Sum('total'))['total'] or 0

    cajeros = Usuario.objects.filter(rol='cajero')

    return render(request, "administrador/reportes/reporte_pedidos_domicilio.html", {
        "pedidos": pedidos,
        "total_pedidos": total_pedidos,
        "total_ventas": total_ventas,
        "cajeros": cajeros,
    })

def obtener_logo():
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        return Image(logo_path, width=110, height=100)
    return None

#-----------------------------
#Exportar en pdf
#-----------------------------

@login_required
@rol_requerido('admin')
def exportar_pedidos_restaurante_excel(request):

    pedidos = (
        Pedido.objects
        .filter(tipo_pedido='restaurante')
        .select_related('mesa', 'mesero')
    )

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pedidos = pedidos.filter(fecha_hora__date__gte=request.GET['inicio'])
    if request.GET.get('fin'):
        pedidos = pedidos.filter(fecha_hora__date__lte=request.GET['fin'])
    if request.GET.get('estado') and request.GET['estado'] != 'todos':
        pedidos = pedidos.filter(estado=request.GET['estado'])
    if request.GET.get('mesero') and request.GET['mesero'] != 'todos':
        pedidos = pedidos.filter(mesero_id=request.GET['mesero'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Restaurante"

    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        logo = ExcelImage(logo_path)
        logo.width = 120
        logo.height = 60
        ws.add_image(logo, "A1")

    # -------- TITULOS --------
    ws.merge_cells("A3:F3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS --------
    headers = ["Código", "Fecha", "Mesa", "Mesero", "Estado", "Total"]
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    for p in pedidos:
        ws.append([
            p.codigo_pedido,
            p.fecha_hora.strftime('%d/%m/%Y %H:%M'),
            p.mesa.numero if p.mesa else "",
            p.mesero.nombre if p.mesero else "",
            p.estado.capitalize(),
            float(p.total)
        ])

    # -------- FORMATO --------
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = center

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pedidos_restaurante.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('admin')
def exportar_pedidos_domicilio_pdf(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio'
    ).select_related('cajero')

    # ===== FILTROS =====
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    estado = request.GET.get("estado")
    cajero = request.GET.get("cajero")

    if inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=inicio)
    if fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fin)
    if estado and estado != "todos":
        pedidos = pedidos.filter(estado=estado)
    if cajero and cajero != "todos":
        pedidos = pedidos.filter(cajero_id=cajero)

    total_final = pedidos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos_domicilio.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    logo = obtener_logo()
    if logo:
        elementos.append(logo)

    elementos.append(Spacer(1, 8))

    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
        ]
    ))

    elementos.append(Spacer(1, 6))
    elementos.append(Table([[f"Fecha de reporte: {fecha_reporte}"]], colWidths=[480]))
    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["REPORTE DE PEDIDOS - DOMICILIO"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 13),
        ]
    ))

    elementos.append(Spacer(1, 10))

    data = [["Pedido", "Cliente", "Cajero", "Estado", "Fecha", "Total"]]

    for p in pedidos:
        cocinero = "—"
        if p.cocinero_listo:
            cocinero = f"{p.cocinero_listo.nombre} {p.cocinero_listo.apellido}"

        data.append([
            p.codigo_pedido,
            p.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            p.mesa.numero if p.mesa else "—",
            p.mesero.nombre if p.mesero else "",
            cocinero,
            p.estado.capitalize(),
            f"$ {p.total:.2f}"
        ])

    tabla = Table(data, colWidths=[55, 85, 40, 85, 95, 55, 55])
    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    elementos.append(Table(
        [["TOTAL GENERAL", f"$ {total_final:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
        ]
    ))

    doc.build(elementos)
    return response

@login_required
@rol_requerido('admin')
def exportar_cobros_domicilio_pdf(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='domicilio',
        estado_pago='confirmado'
    ).select_related('pedido', 'pedido__cajero')

    # ================= FILTROS =================
    if request.GET.get("inicio"):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET["inicio"])

    if request.GET.get("fin"):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET["fin"])

    if request.GET.get("metodo") and request.GET["metodo"] != "todos":
        pagos = pagos.filter(metodo_pago=request.GET["metodo"])

    if request.GET.get("cajero") and request.GET["cajero"] != "todos":
        pagos = pagos.filter(pedido__cajero_id=request.GET["cajero"])
    # ===========================================

    total_final = pagos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cobros_domicilio.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []

    logo = obtener_logo()
    if logo:
        elementos.append(logo)

    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('FONT',(0,0),(-1,-1),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),13),
        ]
    ))

    elementos.append(Spacer(1, 6))
    elementos.append(Table([[f"Fecha de reporte: {fecha_reporte}"]], colWidths=[480]))
    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["REPORTE DE COBROS - DOMICILIO"]],
        colWidths=[480],
        style=[
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor("#cd966c")),
            ('TEXTCOLOR',(0,0),(-1,-1),colors.white),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('FONT',(0,0),(-1,-1),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),14),
        ]
    ))

    elementos.append(Spacer(1, 10))

    data = [["Pedido", "Cliente", "Cajero", "Método", "Total", "Fecha"]]

    for p in pagos:
        data.append([
            p.pedido.codigo_pedido,
            p.pedido.nombre_cliente,
            p.pedido.cajero.nombre if p.pedido.cajero else "",
            p.metodo_pago.capitalize(),
            f"$ {p.total:.2f}",
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[60,120,90,80,60,90])
    tabla.setStyle(TableStyle([
        ('GRID',(0,0),(-1,-1),1,colors.black),
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor("#b4764f")),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONT',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(0,1),(-1,-1),'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    elementos.append(Table(
        [["TOTAL RECAUDADO", f"$ {total_final:.2f}"]],
        colWidths=[350,130],
        style=[
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR',(0,0),(-1,-1),colors.white),
            ('FONT',(0,0),(-1,-1),'Helvetica-Bold'),
            ('ALIGN',(1,0),(1,0),'RIGHT'),
        ]
    ))

    doc.build(elementos)
    return response

@login_required
@rol_requerido('admin')
def exportar_cobros_restaurante_pdf(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='restaurante',
        estado_pago='confirmado'
    ).select_related('pedido', 'pedido__mesero', 'pedido__mesa')

    # ================= FILTROS (IGUALES AL HTML) =================
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    metodo = request.GET.get("metodo")
    mesero_id = request.GET.get("mesero")

    if inicio:
        pagos = pagos.filter(fecha_hora__date__gte=inicio)

    if fin:
        pagos = pagos.filter(fecha_hora__date__lte=fin)

    if metodo and metodo != "todos":
        pagos = pagos.filter(metodo_pago=metodo)

    if mesero_id and mesero_id != "todos":
        pagos = pagos.filter(pedido__mesero_id=mesero_id)

    # ================= TOTALES =================
    total_final = pagos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    # ================= PDF =================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cobros_restaurante.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    # LOGO
    logo = obtener_logo()
    if logo:
        elementos.append(logo)

    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 13),
        ]
    ))

    elementos.append(Spacer(1, 6))

    elementos.append(Table(
        [[f"Fecha de reporte: {fecha_reporte}"]],
        colWidths=[480],
        style=[('ALIGN', (0,0), (-1,-1), 'CENTER')]
    ))

    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["REPORTE DE COBROS - RESTAURANTE"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
        ]
    ))

    elementos.append(Spacer(1, 10))

    # ================= TABLA =================
    data = [["Pedido", "Mesa", "Mesero", "Método", "Total", "Fecha"]]

    for p in pagos:
        data.append([
            p.pedido.codigo_pedido,
            p.pedido.mesa.numero if p.pedido.mesa else "—",
            p.pedido.mesero.nombre if p.pedido.mesero else "",
            p.metodo_pago.capitalize(),
            f"$ {p.total:.2f}",
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[60, 60, 90, 80, 60, 90])
    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    # ================= TOTAL =================
    elementos.append(Table(
        [["TOTAL RECAUDADO", f"$ {total_final:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ]
    ))

    doc.build(elementos)
    return response

#---------------------------
#exportar en excel
#---------------------------
def estilos_tabla():
    header_fill = PatternFill("solid", fgColor="CD966C")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_fill, header_font, center, border

@login_required
@rol_requerido('admin')
def exportar_pedidos_restaurante_pdf(request):

    pedidos = (
        Pedido.objects
        .filter(tipo_pedido='restaurante')
        .select_related('mesa', 'mesero')
    )

    # ======= FILTROS =======
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    estado = request.GET.get("estado")
    mesero = request.GET.get("mesero")

    if inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=inicio)
    if fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=fin)
    if estado and estado != "todos":
        pedidos = pedidos.filter(estado=estado)
    if mesero and mesero != "todos":
        pedidos = pedidos.filter(mesero_id=mesero)

    total_final = pedidos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos_restaurante.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    # LOGO
    logo = obtener_logo()
    if logo:
        elementos.append(logo)

    elementos.append(Spacer(1, 8))

    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
        ]
    ))

    elementos.append(Spacer(1, 6))
    elementos.append(Table([[f"Fecha de reporte: {fecha_reporte}"]], colWidths=[480]))
    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["REPORTE DE PEDIDOS - RESTAURANTE"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 13),
        ]
    ))

    elementos.append(Spacer(1, 10))

    data = [["Pedido", "Fecha", "Mesa", "Mesero", "Estado", "Total"]]

    for p in pedidos:
        data.append([
            p.codigo_pedido,
            p.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            p.mesa.numero if p.mesa else "—",
            p.mesero.nombre if p.mesero else "",
            p.estado.capitalize(),
            f"$ {p.total:.2f}"
        ])

    tabla = Table(data, colWidths=[70, 95, 45, 105, 65, 70])
    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    elementos.append(Table(
        [["TOTAL GENERAL", f"$ {total_final:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ]
    ))

    doc.build(elementos)
    return response

@login_required
@rol_requerido('admin')
def exportar_pedidos_domicilio_excel(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio'
    ).select_related('cajero')

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pedidos = pedidos.filter(fecha_hora__date__gte=request.GET['inicio'])
    if request.GET.get('fin'):
        pedidos = pedidos.filter(fecha_hora__date__lte=request.GET['fin'])
    if request.GET.get('estado') and request.GET['estado'] != 'todos':
        pedidos = pedidos.filter(estado=request.GET['estado'])
    if request.GET.get('cajero') and request.GET['cajero'] != 'todos':
        pedidos = pedidos.filter(cajero_id=request.GET['cajero'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Domicilio"

    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        logo = ExcelImage(logo_path)
        logo.width = 120
        logo.height = 60
        ws.add_image(logo, "A1")

    # -------- TITULOS --------
    ws.merge_cells("A3:G3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:G4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS --------
    headers = ["Código", "Fecha", "Cliente", "Dirección", "Cajero", "Estado", "Total"]
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    for p in pedidos:
        ws.append([
            p.codigo_pedido,
            p.fecha_hora.strftime('%d/%m/%Y %H:%M'),
            p.nombre_cliente,
            p.direccion_entrega,
            p.cajero.nombre if p.cajero else "",
            p.estado.capitalize(),
            float(p.total)
        ])

    # -------- FORMATO --------
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = center

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pedidos_domicilio.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('admin')
def exportar_cobros_restaurante_excel(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='restaurante',
        estado_pago='confirmado'
    ).select_related('pedido', 'pedido__mesero', 'pedido__mesa')

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])

    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])

    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])

    if request.GET.get('cajero') and request.GET['cajero'] != 'todos':
        pagos = pagos.filter(pedido__mesero_id=request.GET['cajero'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobros Restaurante"

    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        logo = ExcelImage(logo_path)
        logo.width = 110
        logo.height = 100
        ws.add_image(logo, "A1")

    # -------- TITULOS --------
    ws.merge_cells("A3:F3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS TABLA --------
    headers = ["Pedido", "Mesa", "Mesero", "Método", "Total", "Fecha"]
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    for p in pagos:
        ws.append([
            p.pedido.codigo_pedido,
            p.pedido.mesa.numero if p.pedido.mesa else "",
            p.pedido.mesero.nombre if p.pedido.mesero else "",
            p.metodo_pago.capitalize(),
            float(p.total),
            p.fecha_hora.strftime('%d/%m/%Y %H:%M')
        ])

    # -------- FORMATO FILAS --------
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # -------- TOTAL --------
    ws.append([])
    total_row = ws.max_row + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="TOTAL RECAUDADO").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=5, value=float(total_recaudado)).font = Font(bold=True)

    # -------- AUTO ANCHO --------
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cobros_restaurante.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('admin')
def exportar_cobros_domicilio_excel(request):

    pagos = Pago.objects.filter(
        pedido__tipo_pedido='domicilio',
        estado_pago='confirmado'
    ).select_related('pedido', 'pedido__cajero')

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])
    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])
    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])
    if request.GET.get('cajero') and request.GET['cajero'] != 'todos':
        pagos = pagos.filter(pedido__cajero_id=request.GET['cajero'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobros Domicilio"

    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 110
        img.height = 100
        ws.add_image(img, "A1")

    # -------- TITULOS --------
    ws.merge_cells("A3:F3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS --------
    headers = ["Pedido", "Cliente", "Cajero", "Método", "Total", "Fecha"]
    ws.append([])
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    for p in pagos:
        ws.append([
            p.pedido.codigo_pedido,
            p.pedido.nombre_cliente,
            p.pedido.cajero.nombre if p.pedido.cajero else "",
            p.metodo_pago.capitalize(),
            float(p.total),
            p.fecha_hora.strftime('%d/%m/%Y %H:%M')
        ])

    # -------- FORMATO --------
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # -------- TOTAL --------
    ws.append([])
    ws.append(["", "", "", "TOTAL RECAUDADO", float(total_recaudado), ""])
    ws[f"E{ws.max_row}"].font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cobros_domicilio.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('admin', 'cajero')
def reporte_unificado(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])

    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])

    if request.GET.get('tipo') and request.GET['tipo'] != 'todos':
        pagos = pagos.filter(pedido__tipo_pedido=request.GET['tipo'])

    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    return render(
        request,
        'administrador/reportes/reporte_unificado.html',
        {
            'pagos': pagos,
            'total_recaudado': total_recaudado
        }
    )

@login_required
@rol_requerido('admin')
def exportar_unificado_pdf(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # ================= FILTROS =================
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    tipo = request.GET.get("tipo")
    metodo = request.GET.get("metodo")

    if inicio:
        pagos = pagos.filter(fecha_hora__date__gte=inicio)

    if fin:
        pagos = pagos.filter(fecha_hora__date__lte=fin)

    if tipo and tipo != "todos":
        pagos = pagos.filter(pedido__tipo_pedido=tipo)

    if metodo and metodo != "todos":
        pagos = pagos.filter(metodo_pago=metodo)

    # ================= TOTALES =================
    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    # ================= PDF =================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_unificado.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    # ================= LOGO =================
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        elementos.append(Image(logo_path, width=110, height=70))

    elementos.append(Spacer(1, 10))

    # ================= NOMBRE LOCAL =================
    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 13),
        ]
    ))

    elementos.append(Spacer(1, 6))

    elementos.append(Table(
        [[f"Fecha de reporte: {fecha_reporte}"]],
        colWidths=[480],
        style=[('ALIGN', (0,0), (-1,-1), 'CENTER')]
    ))

    elementos.append(Spacer(1, 10))

    # ================= TITULO =================
    elementos.append(Table(
        [["REPORTE UNIFICADO - PEDIDOS Y COBROS"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
        ]
    ))

    elementos.append(Spacer(1, 10))

    # ================= TABLA =================
    data = [[
        "Pedido", "Tipo", "Cliente / Mesa",
        "Responsable", "Total", "Método", "Fecha"
    ]]

    for p in pagos:
        pedido = p.pedido
        if not pedido:
            continue  # por si acaso

        # --- Cliente / Mesa ---
        if pedido.tipo_pedido == 'domicilio':
            cliente_mesa = pedido.nombre_cliente or "Domicilio"
        else:
            if pedido.mesa:
                cliente_mesa = f"Mesa {pedido.mesa.numero}"
            else:
                cliente_mesa = "Sin mesa"

        # --- Responsable ---
        if pedido.tipo_pedido == 'domicilio':
            responsable = pedido.cajero.nombre if pedido.cajero else "Cajero no asignado"
        else:
            responsable = pedido.mesero.nombre if pedido.mesero else "Mesero no asignado"

        data.append([
            pedido.codigo_pedido,
            pedido.tipo_pedido.capitalize(),
            cliente_mesa,
            responsable,
            f"$ {p.total:.2f}",
            p.metodo_pago.capitalize(),
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[55, 65, 90, 90, 60, 65, 80])
    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    # ================= TOTAL =================
    elementos.append(Table(
        [["TOTAL RECAUDADO", f"$ {total_recaudado:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ]
    ))

    doc.build(elementos)
    return response

@login_required
@rol_requerido('admin')
def exportar_unificado_excel(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])

    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])

    if request.GET.get('tipo') and request.GET['tipo'] != 'todos':
        pagos = pagos.filter(pedido__tipo_pedido=request.GET['tipo'])

    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Unificado"

    # ===== ESTILOS (IGUAL QUE OTROS REPORTES) =====
    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 110
        img.height = 100
        ws.add_image(img, "A1")

    # -------- TÍTULOS --------
    ws.merge_cells("A3:G3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:G4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS --------
    headers = [
        "Pedido", "Tipo", "Cliente / Mesa",
        "Responsable", "Total", "Método", "Fecha"
    ]
    ws.append([])              # fila vacía
    ws.append(headers)         # encabezados
    header_row = ws.max_row    # normalmente será 6

    # Aplicar estilo de encabezado (como en cobros domicilio)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    data_start_row = header_row + 1

    for p in pagos:
        pedido = p.pedido
        if not pedido:
            continue  # por seguridad

        # ------ Cliente / Mesa ------
        if pedido.tipo_pedido == 'domicilio':
            cliente_mesa = pedido.nombre_cliente or "Domicilio"
        else:
            if pedido.mesa:
                cliente_mesa = f"Mesa {pedido.mesa.numero}"
            else:
                cliente_mesa = "Sin mesa"

        # ------ Responsable ------
        if pedido.tipo_pedido == 'domicilio':
            responsable = pedido.cajero.nombre if pedido.cajero else "Cajero no asignado"
        else:
            responsable = pedido.mesero.nombre if pedido.mesero else "Mesero no asignado"

        # Fecha sin tz para Excel
        fecha_excel = timezone.localtime(p.fecha_hora).replace(tzinfo=None)

        ws.append([
            pedido.codigo_pedido,
            pedido.tipo_pedido.capitalize(),
            cliente_mesa,
            responsable,
            float(p.total),
            p.metodo_pago.capitalize(),
            fecha_excel
        ])

    # Aplicar bordes y centrado a TODA la tabla de datos
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # -------- TOTAL --------
    ws.append([])
    ws.append(["", "", "", "TOTAL RECAUDADO", float(total_recaudado), "", ""])
    ws[f"E{ws.max_row}"].font = Font(bold=True)

    # -------- FORMATO FECHA --------
    for cell in ws["G"]:
        cell.number_format = "DD/MM/YYYY HH:MM"

    # -------- ANCHO COLUMNAS --------
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=\"reporte_unificado.xlsx\"'
    wb.save(response)
    return response

# ======================= REPORTE UNIFICADO CAJERO (WEB) =======================
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_reporte_unificado(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])

    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])

    if request.GET.get('tipo') and request.GET['tipo'] != 'todos':
        pagos = pagos.filter(pedido__tipo_pedido=request.GET['tipo'])

    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    return render(
        request,
        'cajero/reporte/reporte_general.html',
        {
            'pagos': pagos,
            'total_recaudado': total_recaudado
        }
    )

# ======================= REPORTE UNIFICADO CAJERO (PDF) =======================
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_exportar_unificado_pdf(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # ================= FILTROS =================
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    tipo = request.GET.get("tipo")
    metodo = request.GET.get("metodo")

    if inicio:
        pagos = pagos.filter(fecha_hora__date__gte=inicio)

    if fin:
        pagos = pagos.filter(fecha_hora__date__lte=fin)

    if tipo and tipo != "todos":
        pagos = pagos.filter(pedido__tipo_pedido=tipo)

    if metodo and metodo != "todos":
        pagos = pagos.filter(metodo_pago=metodo)

    # ================= TOTALES =================
    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0
    fecha_reporte = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    # ================= PDF =================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_unificado_cajero.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4)

    elementos = []

    # ================= LOGO =================
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        elementos.append(Image(logo_path, width=110, height=70))

    elementos.append(Spacer(1, 10))

    # ================= NOMBRE LOCAL =================
    elementos.append(Table(
        [["CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"]],
        colWidths=[480],
        style=[
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 13),
        ]
    ))

    elementos.append(Spacer(1, 6))

    elementos.append(Table(
        [[f"Fecha de reporte: {fecha_reporte}"]],
        colWidths=[480],
        style=[('ALIGN', (0,0), (-1,-1), 'CENTER')]
    ))

    elementos.append(Spacer(1, 10))

    # ================= TITULO =================
    elementos.append(Table(
        [["REPORTE UNIFICADO - PEDIDOS Y COBROS"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
        ]
    ))

    elementos.append(Spacer(1, 10))

    # ================= TABLA =================
    data = [[
        "Pedido", "Tipo", "Cliente / Mesa",
        "Responsable", "Total", "Método", "Fecha"
    ]]

    for p in pagos:
        pedido = p.pedido
        if not pedido:
            continue  # por si acaso

        # --- Cliente / Mesa ---
        if pedido.tipo_pedido == 'domicilio':
            cliente_mesa = pedido.nombre_cliente or "Domicilio"
        else:
            if pedido.mesa:
                cliente_mesa = f"Mesa {pedido.mesa.numero}"
            else:
                cliente_mesa = "Sin mesa"

        # --- Responsable ---
        if pedido.tipo_pedido == 'domicilio':
            responsable = pedido.cajero.nombre if pedido.cajero else "Cajero no asignado"
        else:
            responsable = pedido.mesero.nombre if pedido.mesero else "Mesero no asignado"

        data.append([
            pedido.codigo_pedido,
            pedido.tipo_pedido.capitalize(),
            cliente_mesa,
            responsable,
            f"$ {p.total:.2f}",
            p.metodo_pago.capitalize(),
            timezone.localtime(p.fecha_hora).strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[55, 65, 90, 90, 60, 65, 80])
    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    # ================= TOTAL =================
    elementos.append(Table(
        [["TOTAL RECAUDADO", f"$ {total_recaudado:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ]
    ))

    doc.build(elementos)
    return response

# ======================= REPORTE UNIFICADO CAJERO (EXCEL) =======================
@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_exportar_unificado_excel(request):

    pagos = Pago.objects.filter(
        estado_pago='confirmado'
    ).select_related(
        'pedido',
        'pedido__mesa',
        'pedido__mesero',
        'pedido__cajero'
    )

    # -------- FILTROS --------
    if request.GET.get('inicio'):
        pagos = pagos.filter(fecha_hora__date__gte=request.GET['inicio'])

    if request.GET.get('fin'):
        pagos = pagos.filter(fecha_hora__date__lte=request.GET['fin'])

    if request.GET.get('tipo') and request.GET['tipo'] != 'todos':
        pagos = pagos.filter(pedido__tipo_pedido=request.GET['tipo'])

    if request.GET.get('metodo') and request.GET['metodo'] != 'todos':
        pagos = pagos.filter(metodo_pago=request.GET['metodo'])

    total_recaudado = pagos.aggregate(total=Sum('total'))['total'] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Unificado"

    # ===== ESTILOS (IGUAL QUE OTROS REPORTES) =====
    header_fill, header_font, center, border = estilos_tabla()

    # -------- LOGO --------
    logo_path = os.path.join(settings.MEDIA_ROOT, 'logo', 'logo.jpeg')
    if os.path.exists(logo_path):
        img = ExcelImage(logo_path)
        img.width = 110
        img.height = 100
        ws.add_image(img, "A1")

    # -------- TÍTULOS --------
    ws.merge_cells("A3:G3")
    ws["A3"] = "CAFÉ RESTAURANTE PRODUCTOS CARLOS GERARDO"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A3"].alignment = center

    ws.merge_cells("A4:G4")
    ws["A4"] = f"Reporte generado: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center

    # -------- ENCABEZADOS --------
    headers = [
        "Pedido", "Tipo", "Cliente / Mesa",
        "Responsable", "Total", "Método", "Fecha"
    ]
    ws.append([])              # fila vacía
    ws.append(headers)         # encabezados
    header_row = ws.max_row    # normalmente será 6

    # Aplicar estilo de encabezado (como en cobros domicilio)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # -------- DATOS --------
    data_start_row = header_row + 1

    for p in pagos:
        pedido = p.pedido
        if not pedido:
            continue  # por seguridad

        # ------ Cliente / Mesa ------
        if pedido.tipo_pedido == 'domicilio':
            cliente_mesa = pedido.nombre_cliente or "Domicilio"
        else:
            if pedido.mesa:
                cliente_mesa = f"Mesa {pedido.mesa.numero}"
            else:
                cliente_mesa = "Sin mesa"

        # ------ Responsable ------
        if pedido.tipo_pedido == 'domicilio':
            responsable = pedido.cajero.nombre if pedido.cajero else "Cajero no asignado"
        else:
            responsable = pedido.mesero.nombre if pedido.mesero else "Mesero no asignado"

        # Fecha sin tz para Excel
        fecha_excel = timezone.localtime(p.fecha_hora).replace(tzinfo=None)

        ws.append([
            pedido.codigo_pedido,
            pedido.tipo_pedido.capitalize(),
            cliente_mesa,
            responsable,
            float(p.total),
            p.metodo_pago.capitalize(),
            fecha_excel
        ])

    # Aplicar bordes y centrado a TODA la tabla de datos
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # -------- TOTAL --------
    ws.append([])
    ws.append(["", "", "", "TOTAL RECAUDADO", float(total_recaudado), "", ""])
    ws[f"E{ws.max_row}"].font = Font(bold=True)

    # -------- FORMATO FECHA --------
    for cell in ws["G"]:
        cell.number_format = "DD/MM/YYYY HH:MM"

    # -------- ANCHO COLUMNAS --------
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=\"reporte_unificado_cajero.xlsx\"'
    wb.save(response)
    return response

#---------------------------------
# REPORTE PEDIDOS CLIENTE
#---------------------------------
@login_required(login_url='login')
@rol_requerido('admin')
def reporte_pedidos_domicilio_cliente(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cliente__isnull=False
    ).select_related(
        'cliente',
        'cajero'
    ).prefetch_related(
        'comprobantes_cliente',
        'pagos__comprobante'
    )

    # ==============================
    # FILTROS
    # ==============================

    codigo = request.GET.get('codigo', '').strip()
    estado = request.GET.get('estado', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    cajero = request.GET.get('cajero', '').strip()

    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if estado:
        pedidos = pedidos.filter(estado=estado)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=parse_date(fecha_inicio))

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=parse_date(fecha_fin))

    if cajero:
        pedidos = pedidos.filter(cajero_id=cajero)

    pedidos = pedidos.order_by('-fecha_hora')

    total_general = pedidos.aggregate(total=Sum('total'))['total'] or 0

    cajeros = Usuario.objects.filter(rol='cajero')

    return render(
        request,
        'administrador/reportes/reporte_pedidos_domicilio_cliente.html',
        {
            'pedidos': pedidos,
            'codigo': codigo,
            'estado': estado,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'cajero_seleccionado': cajero,
            'cajeros': cajeros,
            'total_general': total_general 
        }
    )

@login_required(login_url='login')
@rol_requerido('admin')
def exportar_pedidos_cliente_pdf(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cliente__isnull=False
    ).select_related('cliente', 'cajero')

    # ===== FILTROS =====
    if request.GET.get('codigo'):
        pedidos = pedidos.filter(codigo_pedido__icontains=request.GET['codigo'])

    if request.GET.get('estado'):
        pedidos = pedidos.filter(estado=request.GET['estado'])

    if request.GET.get('fecha_inicio'):
        pedidos = pedidos.filter(fecha_hora__date__gte=request.GET['fecha_inicio'])

    if request.GET.get('fecha_fin'):
        pedidos = pedidos.filter(fecha_hora__date__lte=request.GET['fecha_fin'])

    if request.GET.get('cajero'):
        pedidos = pedidos.filter(cajero_id=request.GET['cajero'])

    pedidos = pedidos.order_by('-fecha_hora')

    total_general = pedidos.aggregate(total=Sum('total'))['total'] or 0

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_pedidos_cliente.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []

    elementos.append(Table(
        [["REPORTE PEDIDOS DOMICILIO CLIENTE"]],
        colWidths=[480],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#cd966c")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
        ]
    ))

    elementos.append(Spacer(1, 10))

    data = [[
        "Código", "Cajero", "Cliente",
        "Total", "Estado", "Fecha"
    ]]

    for p in pedidos:
        cajero = f"{p.cajero.nombre}" if p.cajero else "No asignado"

        data.append([
            p.codigo_pedido,
            cajero,
            f"{p.cliente.nombre}",
            f"$ {p.total:.2f}",
            p.get_estado_display(),
            p.fecha_hora.strftime("%d/%m/%Y %H:%M")
        ])

    tabla = Table(data, colWidths=[70, 80, 80, 60, 80, 90])

    tabla.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b4764f")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,1), (-1,-1), 'CENTER'),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 10))

    elementos.append(Table(
        [["TOTAL GENERAL", f"$ {total_general:.2f}"]],
        colWidths=[350, 130],
        style=[
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#3E7A3F")),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONT', (0,0), (-1,-1), 'Helvetica-Bold'),
        ]
    ))

    doc.build(elementos)
    return response


@login_required(login_url='login')
@rol_requerido('admin')
def exportar_pedidos_cliente_excel(request):

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cliente__isnull=False
    ).select_related('cliente', 'cajero')

    # ===== FILTROS =====
    if request.GET.get('codigo'):
        pedidos = pedidos.filter(codigo_pedido__icontains=request.GET['codigo'])

    if request.GET.get('estado'):
        pedidos = pedidos.filter(estado=request.GET['estado'])

    if request.GET.get('fecha_inicio'):
        pedidos = pedidos.filter(fecha_hora__date__gte=request.GET['fecha_inicio'])

    if request.GET.get('fecha_fin'):
        pedidos = pedidos.filter(fecha_hora__date__lte=request.GET['fecha_fin'])

    if request.GET.get('cajero'):
        pedidos = pedidos.filter(cajero_id=request.GET['cajero'])

    pedidos = pedidos.order_by('-fecha_hora')

    total_general = pedidos.aggregate(total=Sum('total'))['total'] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Cliente"

    headers = ["Código", "Cajero", "Cliente", "Total", "Estado", "Fecha"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for p in pedidos:
        cajero = p.cajero.nombre if p.cajero else "No asignado"

        ws.append([
            p.codigo_pedido,
            cajero,
            p.cliente.nombre,
            float(p.total),
            p.get_estado_display(),
            p.fecha_hora.replace(tzinfo=None)
        ])

    ws.append([])
    ws.append(["", "", "TOTAL GENERAL", float(total_general)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_pedidos_cliente.xlsx"'

    wb.save(response)
    return response


@login_required(login_url='login')
@rol_requerido('cajero')
@never_cache
def tabla_pedidos_pagados_domicilio(request):
    hoy = timezone.localdate()

    pedidos_pagados = (
        Pedido.objects
        .filter(
            tipo_pedido='domicilio',
            cajero=request.user,
            pagos__estado_pago='confirmado',
            pagos__fecha_hora__date=hoy
        )
        .prefetch_related('pagos__comprobante')
        .distinct()
        .order_by('id')
    )

    for p in pedidos_pagados:
        p.pago_confirmado = (
            p.pagos.filter(estado_pago='confirmado')
            .order_by('-id')
            .first()
        )

    html = render_to_string(
        "cajero/pago/tabla_pedidos_pagados_domicilio.html",
        {"pedidos_pagados": pedidos_pagados},
        request=request
    )

    resp = JsonResponse({"html": html})
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    return resp


# ----------------------------
# Cajero - Perfil
# ----------------------------
@login_required(login_url='login')
@rol_requerido('cajero')
def perfil_cajero(request):
    usuario = request.user   # cajero logueado
    return render(request, "cajero/perfil.html", {"usuario": usuario})

@login_required(login_url='login')
@rol_requerido('cajero')
def editar_perfil_cajero(request):
    usuario = request.user  # cajero logueado

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        password = request.POST.get("password", "")
        password2 = request.POST.get("password2", "")

        # -----------------------------
        # VALIDACIONES BACKEND
        # -----------------------------

        # Nombre y apellido: solo letras
        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$', nombre):
            messages.error(request, "El nombre solo debe contener letras.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$', apellido):
            messages.error(request, "El apellido solo debe contener letras.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        # Correo obligatorio + válido
        if not correo:
            messages.error(request, "El correo es obligatorio.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        if not re.match(r"^[^@]+@[^@]+\.[^@]+$", correo):
            messages.error(request, "El formato del correo no es válido.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        # Validar correo no repetido (excepto el mismo usuario)
        if Usuario.objects.exclude(id=usuario.id).filter(correo=correo).exists():
            messages.error(request, "Ya existe un usuario con ese correo.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        # Teléfono obligatorio
        if not telefono:
            messages.error(request, "El teléfono es obligatorio.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        # Teléfono válido
        if not telefono.isdigit():
            messages.error(request, "El teléfono solo debe contener números.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        if len(telefono) != 10:
            messages.error(request, "El teléfono debe tener 10 dígitos.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

        # Dirección obligatoria
        if not direccion:
            messages.error(request, "La dirección es obligatoria.")
            return render(request, "cajero/editar_perfil.html", {"usuario": usuario})
        
        # Contraseñas
        if password or password2:
            # Mínimo 6 caracteres
            if len(password) < 6:
                messages.error(request, "La contraseña debe tener mínimo 6 caracteres.")
                return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

            usuario.password = make_password(password)

        # Guardar datos generales
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.correo = correo
        usuario.username = correo
        usuario.telefono = telefono
        usuario.direccion = direccion
        usuario.save()

        # Actualizar nombre en la sesión (opcional pero recomendable)
        request.session['usuario_nombre'] = f"{usuario.nombre} {usuario.apellido}"

        messages.success(request, "Perfil actualizado correctamente.")
        return redirect("perfil_cajero")

    return render(request, "cajero/editar_perfil.html", {"usuario": usuario})

#-----------------------------
#PERFIL - MESERO
#-----------------------------
@login_required
@rol_requerido('mesero')
def perfil_mesero(request):
    usuario = request.user  # si tu auth usa tu modelo Usuario como user
    return render(request, "mesero/perfil_mesero.html", {"usuario": usuario})

@login_required
@rol_requerido('mesero')
def editar_perfil_mesero(request):
    usuario = request.user

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        correo = request.POST.get("correo", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        direccion = request.POST.get("direccion", "").strip()

        password = request.POST.get("password", "").strip()
        password2 = request.POST.get("password2", "").strip()

        # Actualizar datos
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.correo = correo
        usuario.telefono = telefono
        usuario.direccion = direccion

        # Cambiar contraseña (solo si llenó)
        if password or password2:
            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return redirect("editar_perfil_mesero")

            if len(password) < 6:
                messages.error(request, "La contraseña debe tener mínimo 6 caracteres.")
                return redirect("editar_perfil_mesero")

            usuario.password = make_password(password)

        usuario.save()
        messages.success(request, "Perfil actualizado correctamente.")
        return redirect("perfil_mesero")

    return render(request, "mesero/editar_perfil_mesero.html", {"usuario": usuario})

def service_worker(request):
    sw_path = Path(settings.BASE_DIR) / "Restaurante" / "static" / "pwa" / "sw.js"
    if not sw_path.exists():
        return HttpResponseNotFound("sw.js no encontrado")
    return FileResponse(open(sw_path, "rb"), content_type="application/javascript")


def ping_eliminar(request, notif_id):
    return HttpResponse("ok eliminar " + str(notif_id))

# ----------------------------
# Cliente -Pedido
# ----------------------------

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_listar_pedidos(request):
    hoy = timezone.localdate()

    pedidos = Pedido.objects.filter(
        tipo_pedido='domicilio',
        cliente=request.user,
        fecha_hora__date=hoy
    ).prefetch_related(
        'pagos__comprobante'
    ).order_by('id')

    return render(request, 'cliente/pedidos/listar_pedidos.html', {
        'pedidos': pedidos
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_crear_pedido(request):
    if request.method == 'POST':

        telefono = request.POST.get('telefono', '').strip()
        direccion = request.POST.get('direccion', '').strip()

        if not telefono or not direccion:
            messages.error(request, "Debe completar todos los campos.")
            return redirect('cliente_crear_pedido')

        if not re.fullmatch(r'\d{10}', telefono):
            messages.error(request, "El teléfono debe tener exactamente 10 dígitos numéricos.")
            return redirect('cliente_crear_pedido')

        pedido = Pedido.objects.create(
            cliente=request.user,
            tipo_pedido='domicilio',
            nombre_cliente=f"{request.user.nombre or ''} {request.user.apellido or ''}".strip(),
            contacto_cliente=telefono,
            direccion_entrega=direccion,
            estado='en_creacion'
        )

        return redirect('cliente_agregar_detalles', pedido_id=pedido.id)

    return render(request, 'cliente/pedidos/crear_pedido.html', {
        'cliente': request.user
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_editar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    if request.method == 'POST' and not pedido.detalles.exists():
        messages.error(request, "No puedes guardar un pedido sin productos.")
        return redirect('cliente_editar_pedido', pedido_id=pedido.id)

    if pedido.estado != 'en_creacion':
        messages.error(request, "Solo puedes editar un pedido mientras esté en creación.")
        return redirect('cliente_listar_pedidos')

    if request.method == 'POST':
        telefono = request.POST.get('telefono', '').strip()
        direccion = request.POST.get('direccion', '').strip()

        if not telefono or not direccion:
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('cliente_editar_pedido', pedido_id=pedido.id)

        if not re.fullmatch(r'\d{10}', telefono):
            messages.error(request, "El teléfono debe tener exactamente 10 dígitos numéricos.")
            return redirect('cliente_editar_pedido', pedido_id=pedido.id)

        pedido.contacto_cliente = telefono
        pedido.direccion_entrega = direccion
        pedido.save(update_fields=['contacto_cliente', 'direccion_entrega'])

        messages.success(request, "Pedido actualizado. Revisa tu pago.")
        return redirect('cliente_editar_pago', pedido_id=pedido.id)

    productos = Producto.objects.filter(activo=True).order_by('nombre')

    return render(request, 'cliente/pedidos/editar_pedido.html', {
        'pedido': pedido,
        'productos': productos
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_editar_pago(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    comprobante = pedido.comprobantes_cliente.first()

    if request.method == 'POST':

        numero = request.POST.get('numero_comprobante')
        imagen = request.FILES.get('imagen')

        if not numero:
            messages.error(request, "Debe ingresar el número de comprobante.")
            return redirect('cliente_editar_pago', pedido_id=pedido.id)

        # ===== VALIDAR DUPLICADO EXCLUYENDO EL ACTUAL =====
        qs = ComprobanteCliente.objects.filter(
            numero_comprobante=numero
        )

        if comprobante:
            qs = qs.exclude(id=comprobante.id)

        if qs.exists():
            messages.error(
                request,
                "El número de comprobante ya fue registrado en el sistema."
            )
            return redirect('cliente_editar_pago', pedido_id=pedido.id)
        # =====================================================

        if comprobante:
            comprobante.numero_comprobante = numero
            comprobante.valor = pedido.total
            if imagen:
                comprobante.imagen = imagen
            comprobante.estado = 'pendiente'
            comprobante.save()

        else:
            ComprobanteCliente.objects.create(
                pedido=pedido,
                cliente=request.user,
                numero_comprobante=numero,
                valor=pedido.total,
                imagen=imagen,
                estado='pendiente'
            )

        messages.success(request, "Pago actualizado correctamente.")
        return redirect('cliente_listar_pedidos')

    return render(request, 'cliente/pedidos/editar_pago.html', {
        'pedido': pedido,
        'comprobante': comprobante
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    if pedido.estado != 'en_creacion':
        return JsonResponse({
            "success": False,
            "message": "Solo puede cancelar pedidos en creación."
        })

    if request.method == 'POST':
        pedido.delete()
        return JsonResponse({"success": True})

    return JsonResponse({"success": False})

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_ver_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    comprobante = pedido.comprobantes_cliente.first()

    return render(request, 'cliente/pedidos/ver_pedido.html', {
        'pedido': pedido,
        'comprobante': comprobante
    })

# ----------------------------
# CLIENTE - AGREGAR DETALLES
# ----------------------------

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_agregar_detalles(request, pedido_id):
    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    productos = Producto.objects.filter(activo=True).order_by('nombre')
    detalles = pedido.detalles.order_by('id')

    return render(request, 'cliente/pedidos/agregar_detalles.html', {
        'pedido': pedido,
        'productos': productos,
        'detalles': detalles 
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_agregar_detalle_ajax(request, pedido_id):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))
        observacion = data.get('observacion', '')

        producto = get_object_or_404(Producto, id=producto_id)

        if producto.agotado_hoy:
            return JsonResponse({
                'success': False,
                'mensaje': f'El producto "{producto.nombre}" está agotado hoy.'
            })

        if DetallePedido.objects.filter(pedido_id=pedido_id, producto_id=producto_id).exists():
            return JsonResponse({
                'success': False,
                'mensaje': 'Este producto ya fue agregado.'
            })

        DetallePedido.objects.create(
            pedido_id=pedido_id,
            producto_id=producto_id,
            cantidad=cantidad,
            observacion=observacion
        )

        pedido = Pedido.objects.get(id=pedido_id)
        pedido.recargo = 1.50
        pedido.calcular_totales()

        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string(
            'cliente/pedidos/tabla_detalles.html',
            {'pedido': pedido, 'detalles': detalles},
            request=request
        )

        return JsonResponse({
            'success': True,
            'tabla': tabla_html
        })

    return JsonResponse({'success': False})

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_editar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if pedido.estado != 'en_creacion':
        return JsonResponse({'success': False, 'mensaje': 'No puedes editar este pedido.'})

    if request.method == 'POST':
        import json
        data = json.loads(request.body)

        detalle.cantidad = int(data.get('cantidad', detalle.cantidad))
        detalle.observacion = data.get('observacion', detalle.observacion)
        detalle.save()

        pedido.calcular_totales()
        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string(
            'cliente/pedidos/tabla_detalles.html',
            {'pedido': pedido, 'detalles': detalles},
            request=request
        )

        return JsonResponse({'success': True, 'tabla': tabla_html})

    return JsonResponse({'success': False})

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_eliminar_detalle_ajax(request, detalle_id):
    detalle = get_object_or_404(DetallePedido, id=detalle_id)
    pedido = detalle.pedido

    if pedido.estado != 'en_creacion':
        return JsonResponse({'success': False, 'mensaje': 'No puedes eliminar este pedido.'})

    if request.method == 'POST':
        detalle.delete()
        pedido.calcular_totales()

        detalles = pedido.detalles.order_by('id')

        tabla_html = render_to_string(
            'cliente/pedidos/tabla_detalles.html',
            {'pedido': pedido, 'detalles': detalles},
            request=request
        )

        return JsonResponse({'success': True, 'tabla': tabla_html})

    return JsonResponse({'success': False})

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_pago_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio'
    )

    if not pedido.detalles.exists():
        messages.error(request, "No puedes pagar un pedido vacío.")
        return redirect('cliente_agregar_detalles', pedido_id=pedido.id)

    if pedido.comprobantes_cliente.exists():
        messages.warning(request, "Ya enviaste un comprobante para este pedido.")
        return redirect('cliente_listar_pedidos')

    if request.method == 'POST':

        imagen = request.FILES.get('imagen')
        numero = request.POST.get('numero_comprobante')

        print("FILES:", request.FILES)
        print("IMAGEN:", imagen)

        if not imagen or not numero:
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect(request.path)

        #VALIDAR DUPLICADO (primera barrera)
        if ComprobanteCliente.objects.filter(
            numero_comprobante=numero
        ).exists():
            messages.error(
                request,
                "El número de comprobante ya fue registrado en el sistema."
            )
            return redirect(request.path)

        try:
            #CREAR (segunda barrera con try/except)
            ComprobanteCliente.objects.create(
                pedido=pedido,
                cliente=request.user,
                numero_comprobante=numero,
                valor=pedido.total,
                imagen=imagen,
                estado='pendiente'
            )

        except IntegrityError:
            messages.error(
                request,
                "El número de comprobante ya fue registrado en el sistema."
            )
            return redirect(request.path)
        
        pedido.estado = 'pendiente_caja'
        pedido.cajero = None
        pedido.enviado_cocina = False
        pedido.save(update_fields=['estado','cajero','enviado_cocina'])

        channel_layer = get_channel_layer()

        comp = ComprobanteCliente.objects.filter(pedido=pedido).first()

        async_to_sync(channel_layer.group_send)(
            "pedidos_cajero",
            {
                "type": "nuevo_pedido_cajero",
                "pedido": {
                    "id": pedido.id,
                    "codigo_pedido": pedido.codigo_pedido,
                    "nombre_cliente": pedido.cliente.nombre,
                    "direccion": pedido.direccion_entrega,
                    "total": float(pedido.total),
                    "comprobante_numero": comp.numero_comprobante,
                    "comprobante_url": comp.imagen.url if comp and comp.imagen else ""
                }
            }
        )

        # Notificar cajeros
        cajeros = Usuario.objects.filter(rol='cajero')

        for cajero in cajeros:
            Notificacion.objects.create(
                pedido=pedido,
                usuario_destino=cajero,
                tipo='pago_pendiente',
                mensaje=f'Nuevo comprobante de pago del cliente {request.user.nombre}'
            )

            enviar_push(
                cajero,
                "Nuevo pedido pendiente",
                f"Nuevo comprobante del cliente {request.user.nombre}",
                "/cajero/pedidos-clientes-domicilio/"
            )

        messages.success(request, "Comprobante enviado correctamente.")
        return redirect('cliente_listar_pedidos')

    return render(request, 'cliente/pedidos/pago.html', {
        'pedido': pedido
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_enviar_pedido(request, pedido_id):

    pedido = get_object_or_404(
        Pedido,
        id=pedido_id,
        cliente=request.user,
        tipo_pedido='domicilio',
        estado='en_creacion'
    )

    if not pedido.detalles.exists():
        return JsonResponse({
            "success": False,
            "message": "No puedes enviar un pedido vacío."
        })

    pedido.estado = 'pendiente_caja'
    pedido.cajero = None          # ← AQUI
    pedido.enviado_cocina = False # ← YA LO TENÍAS
    pedido.save(update_fields=['estado','enviado_cocina','cajero'])


    channel_layer = get_channel_layer()

    cajeros = Usuario.objects.filter(rol='cajero')

    for cajero in cajeros:

        # guardar notificación DB
        notif = Notificacion.objects.create(
            pedido=pedido,
            usuario_destino=cajero,
            tipo='pedido_cliente',
            mensaje=f'Nuevo pedido a domicilio #{pedido.codigo_pedido}'
        )

        # ENVÍO REAL TIME
        async_to_sync(channel_layer.group_send)(
            f"notificaciones_{cajero.id}",
            {
                "type": "enviar_notificacion",
                "tipo": "pedido_cliente",
                "mensaje": f"Nuevo pedido a domicilio por revisar",
                "pedido": pedido.id,
                "codigo_pedido": pedido.codigo_pedido,
                "id": notif.id,
                "fecha": str(notif.fecha)
            }
        )

        enviar_push(
            cajero,
            "Nuevo pedido",
            f"Nuevo pedido #{pedido.codigo_pedido}",
            "/cajero/pedidos-clientes-domicilio/"
        )

    return JsonResponse({"success": True})

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_historial_pedidos(request):

    codigo = request.GET.get('codigo', '').strip()
    estado = request.GET.get('estado', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()

    pedidos = Pedido.objects.filter(
        cliente=request.user,
        tipo_pedido='domicilio',
        estado__in=['listo', 'rechazado']
    ).prefetch_related(
        'comprobantes_cliente'
    )

    if codigo:
        pedidos = pedidos.filter(codigo_pedido__icontains=codigo)

    if estado in ['listo', 'rechazado']:
        pedidos = pedidos.filter(estado=estado)

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_hora__date__gte=parse_date(fecha_inicio))

    if fecha_fin:
        pedidos = pedidos.filter(fecha_hora__date__lte=parse_date(fecha_fin))

    pedidos = pedidos.order_by('-fecha_hora')

    return render(request, 'cliente/pedidos/historial_pedidos.html', {
        'pedidos': pedidos,
        'codigo': codigo,
        'estado': estado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_ver_perfil(request):
    return render(request, 'cliente/perfil/ver_perfil.html', {
        'usuario': request.user
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_editar_perfil(request):

    usuario = request.user

    if request.method == 'POST':

        usuario.nombre = request.POST.get('nombre', '').strip()
        usuario.apellido = request.POST.get('apellido', '').strip()
        usuario.correo = request.POST.get('correo', '').strip()
        usuario.telefono = request.POST.get('telefono', '').strip()
        usuario.direccion = request.POST.get('direccion', '').strip()

        password = request.POST.get('password', '').strip()
        password2 = request.POST.get('password2', '').strip()

        # ===== VALIDAR CONTRASEÑAS =====
        if password or password2:
            if password != password2:
                messages.error(request, "Las contraseñas no coinciden.")
                return redirect('cliente_editar_perfil')

            if len(password) < 6:
                messages.error(request, "La contraseña debe tener al menos 6 caracteres.")
                return redirect('cliente_editar_perfil')

            usuario.set_password(password)

        try:
            usuario.save()
        except IntegrityError:
            messages.error(request, "El correo ingresado ya está registrado.")
            return redirect('cliente_editar_perfil')

        messages.success(request, "Perfil actualizado correctamente.")
        return redirect('cliente_ver_perfil')

    return render(request, 'cliente/perfil/editar_perfil.html', {
        'usuario': usuario
    })

@login_required(login_url='login')
@rol_requerido('cliente')
def cliente_desactivar_cuenta(request):
    usuario = request.user

    # Seguridad extra
    if usuario.rol != 'cliente':
        messages.error(request, "No tienes permiso para realizar esta acción.")
        return redirect('login')

    # ¿Tiene pedidos?
    tiene_pedidos = usuario.pedidos_cliente.exists()

    if tiene_pedidos:
        # SOLO DESACTIVAR
        usuario.is_active = False
        usuario.save(update_fields=['is_active'])

        messages.info(
            request,
            "Tu cuenta ha sido desactivada."
        )

    else:
        # ELIMINAR DEFINITIVAMENTE
        usuario.delete()

        messages.success(
            request,
            "Tu cuenta fue eliminada definitivamente del sistema."
        )

    # Cerrar sesión siempre
    auth_logout(request)
    request.session.flush()

    return redirect('login')

def cliente_reactivar_cuenta(request):
    if request.method == 'POST':
        correo = request.POST.get('correo', '').strip().lower()

        try:
            usuario = Usuario.objects.get(
                correo=correo,
                rol='cliente',
                is_active=False
            )
        except Usuario.DoesNotExist:
            messages.error(
                request,
                "No existe una cuenta desactivada con ese correo."
            )
            return redirect('cliente_reactivar_cuenta')

        uid = urlsafe_base64_encode(force_bytes(usuario.pk))
        token = default_token_generator.make_token(usuario)

        enlace = request.build_absolute_uri(
            reverse('cliente_confirmar_reactivacion', args=[uid, token])
        )

        send_mail(
            subject="Reactivar cuenta - Café Restaurante",
            message=(
                f"Hola {usuario.nombre},\n\n"
                f"Solicitaste reactivar tu cuenta.\n\n"
                f"Haz clic en el siguiente enlace:\n\n"
                f"{enlace}\n\n"
                f"Si no fuiste tú, ignora este mensaje."
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[correo],
            fail_silently=False,
        )

        messages.success(
            request,
            "Te enviamos un enlace para reactivar tu cuenta."
        )
        return redirect('login')

    return render(request, 'login/cliente_reactivar_cuenta.html')


def cliente_confirmar_reactivacion(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        usuario = Usuario.objects.get(pk=uid, rol='cliente')
    except Exception:
        usuario = None

    if usuario is None or not default_token_generator.check_token(usuario, token):
        messages.error(
            request,
            "El enlace de reactivación no es válido o ya expiró."
        )
        return redirect('login')

    usuario.is_active = True
    usuario.save(update_fields=['is_active'])

    messages.success(
        request,
        "Tu cuenta ha sido reactivada correctamente. Ya puedes iniciar sesión."
    )
    return redirect('login')

@login_required(login_url='login')
@rol_requerido('cajero')
def cajero_reportes(request):
    return render(request, 'cajero/reporte/reportes.html')

#-----------------
#PUSH
#-----------------

@csrf_exempt
@login_required
def guardar_suscripcion(request):
    data = json.loads(request.body)

    PushSubscription.objects.update_or_create(
        endpoint=data["endpoint"],
        defaults={
            "usuario": request.user,
            "p256dh": data["keys"]["p256dh"],
            "auth": data["keys"]["auth"],
        }
    )

    return JsonResponse({"ok": True})